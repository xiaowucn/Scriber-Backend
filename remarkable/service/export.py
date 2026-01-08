import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from utensils.hash import md5sum

from remarkable.common.constants import HistoryAction
from remarkable.common.enums import AuditAnswerType
from remarkable.common.exceptions import NoModelFoundError
from remarkable.common.storage import localstorage
from remarkable.common.util import clean_txt, dump_data_to_worksheet, remove_illegal_characters
from remarkable.config import project_root
from remarkable.converter import SimpleJSONConverter, flatten_dict
from remarkable.db import pw_db
from remarkable.models.new_file import NewFile
from remarkable.optools.export_answers_for_szse import fetch_all_answer_data
from remarkable.pdfinsight.reader import PdfinsightReader
from remarkable.pw_models.audit_rule import NewAuditResult
from remarkable.pw_models.model import NewMold
from remarkable.pw_models.question import NewQuestion


class Exporter:
    def __init__(
        self,
        task_id,
        project_id=None,
        mold_id=None,
        tree_l=None,
        training_data=None,
        files_ids=None,
        headers=None,
        export_action=None,
    ):
        self.task_id = task_id
        self.project_id = project_id
        self.mold_id = mold_id
        self.tree_l = tree_l
        self.training_data = training_data
        self.files_ids = files_ids
        self.headers = headers
        self.export_action = export_action

    async def get_all_questions(self):
        questions = []
        if self.project_id:
            questions = await NewQuestion.list_by_range(project=self.project_id, special_cols=["id", "fid", "answer"])
        elif self.tree_l:
            questions = await NewQuestion.list_by_range(tree_l=self.tree_l, special_cols=["id", "fid", "answer"])
        elif self.files_ids:
            questions = await NewQuestion.list_by_range(files_ids=self.files_ids, special_cols=["id", "fid", "answer"])
        return questions

    async def get_all_files(self):
        files = []
        if self.project_id:
            files = await NewFile.find_by_pid(pid=self.project_id)
        elif self.tree_l:
            files = await NewFile.list_by_range(tree_l=self.tree_l)
        elif self.files_ids:
            files = await NewFile.find_by_ids(self.files_ids)
        return files


class GfFundExporter(Exporter):
    async def get_files_table_of_content(self):
        files = await self.get_all_files()
        first_level_titles = []
        second_level_titles = []
        for file in files:
            if pdfinsight_path := file.pdfinsight_path():
                pdfinsight = PdfinsightReader(localstorage.mount(pdfinsight_path))
                first_level_directories = []
                second_level_directories: list[dict] = []
                for syllabus in pdfinsight.syllabus_dict.values():
                    if syllabus["level"] == 1:
                        first_level_directories.append(syllabus["title"])
                        # https://gitpd.paodingai.com/cheftin/docs_scriber/-/issues/1594#note_295516
                        if syllabus["children"] and "重要提示" not in syllabus["title"]:
                            second_level_directories.append(
                                {
                                    "parent": syllabus["title"],
                                    "children": [
                                        pdfinsight.syllabus_dict[child]["title"] for child in syllabus["children"]
                                    ],
                                }
                            )
                            second_level_directories[-1].update(
                                {
                                    "categorical_value": md5sum(
                                        clean_txt("".join(second_level_directories[-1]["children"]))
                                    )
                                }
                            )
                titles_md5 = [content["categorical_value"] for content in second_level_directories]
                categorical_value = md5sum(clean_txt("".join(first_level_directories)))
                titles_md5.append(categorical_value)
                directory_characteristic_value = md5sum("".join(titles_md5))
                # 文件ID 文件名 目录特征值	目录类型	分类值 所属一级 章节标题
                first_level_titles.append(
                    [
                        file.id,
                        file.name,
                        directory_characteristic_value,
                        "一级",
                        categorical_value,
                        "-1",
                        *first_level_directories,
                    ]
                )
                for second_title in second_level_directories:
                    second_level_titles.append(
                        [
                            file.id,
                            file.name,
                            directory_characteristic_value,
                            "二级",
                            second_title["categorical_value"],
                            second_title["parent"],
                            *second_title["children"],
                        ]
                    )
        return first_level_titles + second_level_titles

    async def export(self):
        if self.export_action == HistoryAction.CREATE_TABLE_OF_CONTENT:
            if not (table_of_content := await self.get_files_table_of_content()):
                return None
            header = ["文件ID", "文件名", "目录特征值", "目录类型", "分类值", "所属一级"]
            workbook = openpyxl.Workbook(write_only=True)
            dump_data_to_worksheet(workbook, header, data=table_of_content)
        else:
            questions = await self.get_all_questions()
            if not questions:
                return None
            workbook = openpyxl.Workbook(write_only=True)
            mold = await NewMold.find_by_id(self.mold_id)
            if not mold:
                raise NoModelFoundError(msg="mold not found")
            user_answers = {question.fid: question.answer for question in questions}
            converted_answer = {
                fid: SimpleJSONConverter(user_answer).convert(item_handler=lambda x: x.origin_text)
                for fid, user_answer in user_answers.items()
                if user_answer
            }
            self.dump_data_to_worksheet(workbook, converted_answer)
        root_path = Path(project_root)
        dump_filename = root_path / "data" / "export_answer_data" / f"task_{self.training_data.id}.xlsx"
        workbook.save(dump_filename)
        return dump_filename

    @staticmethod
    def group_answer(converted_answer):
        """
        分组分组合类型和普通类型答案
        """
        normal_answer = []
        combination_answer = []
        for fid, answer in converted_answer.items():
            combination_type = {
                "文件ID": str(fid),
                "基金名称": answer.get("基金名称"),
                "基金代码": answer.get("基金代码"),
                "报告名称": answer.get("报告名称"),
                "基金全称": answer.get("基金全称"),
            }
            normal_type = {"文件ID": str(fid)}
            for key, value in answer.items():
                if isinstance(value, list):
                    combination_type[key] = value
                else:
                    normal_type[key] = value
            combination_answer.append(combination_type)
            normal_answer.append(normal_type)
        return combination_answer, normal_answer

    def dump_data_to_worksheet(self, workbook, converted_answer):
        """
        将组合类型的答案和普通一对一的答案分别保存在不同的sheet中
        """
        combination_answer, normal_answer = self.group_answer(converted_answer)
        # 处理普通一对一的答案
        normal_sheet = workbook.create_sheet("份额共有信息", 0)
        normal_header = list(normal_answer[0].keys())
        normal_sheet.append(normal_header)
        normal_data = [list(row_data.values()) for row_data in normal_answer]
        for row_data in normal_data:
            normal_sheet.append(row_data)
        convert_sheet_name = {
            "基金业绩表现": "业绩文段数据",
            "报告期时间": "业绩图始末时间",
        }
        reverse_dict = {v: k for k, v in convert_sheet_name.items()}
        # 处理组合类型的答案
        combination_sheets = []
        for key, value in list(converted_answer.values())[0].items():
            if isinstance(value, list):
                key = convert_sheet_name.get(key, key)
                combination_sheets.append(workbook.create_sheet(key))

        for sheet in combination_sheets:
            sheet_header = []
            for answer in combination_answer:
                title = reverse_dict.get(sheet.title, sheet.title)
                combination_data = answer[title]
                if combination_data:
                    if not sheet_header:
                        # group_answer() 里添加的combination_type里的答案
                        sheet_header += [k for k, v in answer.items() if isinstance(v, str)]

                        sheet_header.extend(self.combination_dict_key(title, combination_data[0]))
                        sheet.append(sheet_header)
                    cells = [v for _, v in answer.items() if isinstance(v, str)]
                    for cell in combination_data:
                        row_data = cells[:] + list(cell.values())
                        sheet.append(row_data)

    @staticmethod
    def combination_dict_key(title, data):
        """
        title: 报告期时间
        data: {'份额名称': '广发百发大数据精选混合A', '期初时间': '2015-09-14', '期末时间': '2022-03-31'}
        return ['报告期时间-份额名称', '报告期时间-期初时间', '报告期时间-期末时间']
        """
        return [f"{title}-{key}" for key in data]


class InspectDataExporter(Exporter):
    async def export(self):
        data = []
        header = ["文档ID", "文档名称", "规则名称", "字段名称", "原文内容", "合规结果", "不通过原因", "修改意见"]
        root_path = Path(project_root)
        # dump_filename = root_path / "data" / "export_answer_data" / f'task_{self.training_data.id}.xlsx'
        dump_filename = root_path / "data" / "export_answer_data" / f"task_{self.mold_id}.xlsx"

        files = await self.get_all_files()
        for file in files:
            cond = (
                (NewAuditResult.fid == file.id)
                & (NewAuditResult.schema_id == self.mold_id)
                & (NewAuditResult.answer_type == AuditAnswerType.final_answer)
            )
            inspect_items = await pw_db.execute(NewAuditResult.select().where(cond))
            for item in inspect_items:
                data.append(
                    (
                        file.id,
                        file.name,
                        item.name,
                        "\n".join(x["name"] for x in item.schema_results) if item.schema_results else "",
                        "\n".join(f"{x['name']}: {x.get('text', '')}" for x in item.schema_results)
                        if item.schema_results
                        else "",
                        "合规" if item.is_compliance else "不合规",
                        "\n".join(x["reason_text"] for x in item.reasons),
                        item.suggestion,
                    )
                )
        workbook = openpyxl.Workbook()
        dump_data_to_worksheet(workbook, header, data)
        self.optimize_xls_style(workbook)
        workbook.save(dump_filename)
        return dump_filename

    @staticmethod
    def optimize_xls_style(workbook):
        width_maps = {
            "A": 10,
            "B": 50,
            "C": 50,
            "D": 40,
            "E": 45,
            "F": 10,
            "G": 40,
            "H": 40,
        }
        worksheet = workbook.active
        for col_index, width in width_maps.items():
            worksheet.column_dimensions[col_index].width = width
            if col_index in ("D", "E"):
                horizontal = "left"
            else:
                horizontal = "center"
            for col in worksheet.iter_cols():
                for cell in col:
                    cell.font = Font(size=14)
                    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


class CmfChinaExporter(Exporter):
    async def export(self) -> str:
        mold = await NewMold.find_by_id(self.mold_id)
        if not mold:
            raise NoModelFoundError(msg="mold not found")
        answer_data_list = await fetch_all_answer_data(mold.id, files_ids=self.files_ids)

        workbook = openpyxl.Workbook()
        self.dump_data_to_worksheet(workbook, answer_data_list, mold.name)
        self.optimize_xls_style(workbook, mold.name)
        data_path = os.path.join(project_root, "data", "export_answer_data")
        os.makedirs(data_path, exist_ok=True)
        dump_filename = os.path.join(data_path, f"task_{self.training_data.id}.xlsx")
        workbook.save(dump_filename)
        return dump_filename

    def dump_data_to_worksheet(self, workbook, answer_data, sheet_name: str = "sheet", sheet_index=0):
        sheet = workbook.create_sheet(sheet_name, sheet_index)
        sheet.append(["文档ID", "文档名称", "字段路径", "字段值"])
        for record in answer_data:
            for key, value in flatten_dict(SimpleJSONConverter(record.get("data", {})).convert(), keep_index=True):
                sheet.append(
                    remove_illegal_characters(
                        [
                            record["fid"],
                            record["name"],
                            key,
                            value,
                        ]
                    )
                )

    @staticmethod
    def optimize_xls_style(workbook, sheet_name: str = "sheet"):
        width_maps = {
            "A": 10,
            "B": 50,
            "C": 50,
            "D": 50,
        }
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        for col_index, width in width_maps.items():
            worksheet.column_dimensions[col_index].width = width
            for col in worksheet.iter_cols():
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
