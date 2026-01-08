import asyncio
import logging
import tempfile
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from remarkable.answer.reader import AnswerReader
from remarkable.common.constants import QuestionStatus
from remarkable.common.storage import localstorage
from remarkable.common.util import box_to_outline
from remarkable.config import target_path
from remarkable.db import pw_db
from remarkable.models.new_file import NewFile
from remarkable.optools.fm_upload import FMUploader
from remarkable.pdfinsight.reader import PdfinsightReader
from remarkable.pw_models.question import NewQuestion

TEMPLATE_PATH: str = target_path("data/htsc/template.xlsx")


class ExcelParser:
    SHEETS = ["原始信息", "段落", "关键词", "问句"]
    HEADER = ["股票代码", "股票简称", "二级行业名称", "财报标题", "s3_url", "doc_id", "BGEdoc_id"]
    HEADER_DICT = {
        "段落": ["摘要段落内容", "摘要段落的核心（人工标注）"],
        "关键词集合": ["关键词", "核心"],
        "问句": ["问句", "问句内容"],
    }
    FONT = Font(bold=True)
    ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def __init__(self, sheet_name: str):
        self.data = {}
        self.sheet_name = sheet_name
        self.max_header = {}
        self.data_list = []

    def load_excel(self, sheet_name: str):
        workbook = load_workbook(TEMPLATE_PATH)
        sheet = workbook[sheet_name]
        self.data[sheet_name] = ExcelParser.load_sheet(sheet)

    @classmethod
    def load_sheet(cls, sheet: Worksheet):
        res = {}
        for row in sheet.iter_rows(min_row=2):
            row_data = [c.value if c.value is not None else "" for c in row]
            doc_id = row_data[5].strip()
            if doc_id:
                res[doc_id] = row_data
        return res

    def format_data(self, fid, answer, pdfinsight_path):
        max_answer = {
            "段落": [0, []],
            "关键词集合": [0, []],
            "问句": [0, []],
        }

        reader = PdfinsightReader(localstorage.mount(pdfinsight_path))
        _root, _mapping = AnswerReader(answer).build_answer_tree()
        answers = _root.to_dict()
        if not answers["华泰测试集"]:
            logging.warning(f"{fid} has no answer")
            return
        answers_dict = answers["华泰测试集"][0]
        doc_id = answers_dict["文档ID"]["data"][0]["text"].strip()
        template_field = self.data[self.sheet_name].get(doc_id)
        fixed_value = template_field[:7] if template_field else ""
        answer_list = answers_dict["问题"]
        for item in answer_list:
            fixed_value_, q_answer, q_type = self.get_question_info(item)
            q_list = []
            same_answer_index_dict = self.get_same_answer_index(q_answer)
            self.get_answer_data(same_answer_index_dict, q_answer, q_list, reader)
            q_type_answer = max_answer[q_type]
            q_type_answer[1].append(fixed_value_ + q_list)
            length = self._count_length(same_answer_index_dict)
            q_type_answer[0] = len(q_answer) - length if q_type_answer[0] < len(q_answer) - length else q_type_answer[0]
        if fixed_value:
            self.data_list.append((fixed_value, max_answer))

    @staticmethod
    def _count_length(same_answer_index):
        length = 0
        for v in same_answer_index.values():
            if len(v) > 1:
                length += len(v) - 1
        return length

    @staticmethod
    def get_question_info(item: dict):
        q_type = item.get("问题类型").plain_text if item.get("问题类型") else "段落"
        q_content = item.get("问题内容").plain_text
        q_core = item.get("问题核心").plain_text
        q_answer = item.get("答案")
        fixed_value_ = [q_content, q_core]
        return fixed_value_, q_answer, q_type

    def writer_excel(self):
        workbook = Workbook()
        del workbook[workbook.sheetnames[0]]
        for fixed_value, max_answer_lengths in self.data_list:
            for k, v in max_answer_lengths.items():
                if not v[1]:
                    continue
                sheet = workbook[k] if k in workbook.sheetnames else workbook.create_sheet(title=k)
                headers = self.HEADER + self.HEADER_DICT[k]
                headers += [
                    f"召回段落{num}{extra}（人工标注）"
                    for num in range(1, v[0] + 1)
                    for extra in ["内容", "index", "页码", "类型", "核心"]
                ]
                self.max_header[k] = headers if len(self.max_header.get(k, "")) < len(headers) else self.max_header[k]
                for values in v[1]:
                    sheet.append(fixed_value + values)

        for k, v in self.max_header.items():
            ws = workbook[k]
            ws.insert_rows(idx=1)
            for index, i in enumerate(v):
                ws.cell(row=1, column=index + 1).value = i

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = self.ALIGNMENT

        with tempfile.NamedTemporaryFile(prefix="htsc-test-set", suffix=".xlsx") as tmp_fp:
            excel_path = tmp_fp.name
            workbook.save(excel_path)
            FMUploader().upload(Path(excel_path))

    @staticmethod
    def get_answer_data(
        same_answer_index: dict[str, list[int]], q_answer: list[dict], q_list: list, reader: PdfinsightReader
    ):
        skip_index = []
        for index, answer in enumerate(q_answer):
            if index in skip_index:
                continue

            answer_content = answer.get("答案内容").plain_text if answer.get("答案内容") else ""
            answer_box_list = answer["答案内容"]["data"] if answer.get("答案内容") else []
            if not answer.get("答案类型"):
                all_elements = []
                for answer_box in answer_box_list:
                    page = answer_box["boxes"][0]["page"]
                    outline = box_to_outline(answer_box["boxes"][0]["box"])
                    elements = reader.find_elements_by_outline(page, outline)
                    all_elements.extend(elements)
                answer_type = "表格" if any(ele.get("class") == "TABLE" for _, ele in all_elements) else "段落"
            else:
                answer_type = answer.get("答案类型").plain_text

            same_index_list = same_answer_index.get(answer_content + str(answer_box_list))
            answer_core_list = []
            for idx in same_index_list:
                item = q_answer[idx]
                answer_core_text = item["答案核心"].plain_text if item.get("答案核心") else ""
                answer_core_list.append(answer_core_text)
                if idx != index:
                    skip_index.append(idx)
            answer_core = "、".join(set(answer_core_list))

            answers_pages, answers_indexes = [], []
            for answer_box in answer_box_list:
                page = answer_box["boxes"][0]["page"]
                outline = box_to_outline(answer_box["boxes"][0]["box"])
                etype, element = reader.find_element_by_outline(page, outline)
                if not element:
                    continue
                answers_pages.append(str(element["page"]))
                answers_indexes.append(str(element["index"]))
            answer_page = "、".join(answers_pages)
            answer_index = "、".join(answers_indexes)
            q_list.extend([answer_content, answer_index, answer_page, answer_type, answer_core])

    @staticmethod
    def get_same_answer_index(q_answer: list[dict]) -> dict[str, list[int]]:
        same_answer_index_dict = {}
        same_answer_dict = {}
        for i, answer in enumerate(q_answer):
            answer_content = answer["答案内容"]
            answer_key = (answer_content.plain_text + str(answer_content["data"])) if answer_content else ""
            if answer_key in same_answer_dict:
                same_answer_dict[answer_key].append(i)
            else:
                same_answer_dict[answer_key] = [i]
        for v in same_answer_dict.values():
            if len(v) > 1:
                same_answer_index_dict[v[0]] = v[1:]
        return same_answer_dict


async def get_answer_and_pdf_path(fid: int) -> tuple:
    question = await NewQuestion.find_by_fid(fid)
    if not question:
        return "", ""

    question = question[0]
    file_info = await NewFile.find_by_qid(question.id)
    pdfinsight_path = NewFile.get_path(file_info.pdfinsight)

    return question.answer, pdfinsight_path


async def get_fids_by_pid(pid):
    cond = NewQuestion.status == QuestionStatus.FINISH.value
    cond &= NewFile.pid == pid
    query = NewFile.select(NewFile.id).join(NewQuestion, on=(NewFile.id == NewQuestion.fid)).where(cond)
    return await pw_db.scalars(query)


async def main(sheet_name: str = "测试集原始信息", pid: int = 123, test_fid=None):
    logging.info("starting...")
    excel = ExcelParser(sheet_name)
    excel.load_excel(sheet_name)
    if test_fid:
        fids = [test_fid]
    else:
        fids = await get_fids_by_pid(pid)
    for fid in fids:
        logging.info(f"fid: {fid}")
        answer, pdfinsight_path = await get_answer_and_pdf_path(fid)
        if answer and pdfinsight_path:
            excel.format_data(fid, answer, pdfinsight_path)
    if excel.data_list:
        excel.writer_excel()
    logging.info("end")


if __name__ == "__main__":
    # asyncio.run(main(test_fid =2318))
    asyncio.run(main())
