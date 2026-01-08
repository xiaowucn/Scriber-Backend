# pylint: skip-file
import json
import logging
import re
import shutil
import urllib
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from functools import cached_property
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from remarkable.answer.node import AnswerItem
from remarkable.common.exceptions import CustomError
from remarkable.common.pattern import PatternCollection
from remarkable.common.schema import Schema, attribute_id
from remarkable.common.storage import localstorage, tmp_storage
from remarkable.common.util import clean_txt, dump_data_to_worksheet
from remarkable.config import get_config
from remarkable.db import pw_db
from remarkable.models.model_version import NewModelVersion
from remarkable.models.new_file import NewFile
from remarkable.pdfinsight.reader import PdfinsightReader
from remarkable.plugins.fileapi.financial_attribute import FinancialAttribute
from remarkable.pw_models.model import NewSpecialAnswer


@dataclass
class Url:
    fid: int
    mid: int
    qid: int
    tree_id: int
    schema_key: str
    scheme: str = get_config("web.scheme", "http")
    domain: str = get_config("web.domain", "localhost:8000")

    def __str__(self):
        if self.schema_key.endswith("-原文"):
            self.schema_key = self.schema_key.rstrip("-原文")
        if "-" in self.schema_key:
            self.schema_key = self.schema_key.split("-")[-1]
        self.schema_key = urllib.parse.quote(self.schema_key)
        return f"{self.scheme}://{self.domain}/#/project/remark/{self.qid}?projectId={self.tree_id}&treeId={self.tree_id}&fileId={self.fid}&schemaId={self.mid}&schemaKey={self.schema_key}"


class ErrorReport:
    recall_errors = defaultdict(set)
    precision_errors = defaultdict(set)
    fit_stats = defaultdict(set)
    export_data = defaultdict(list)

    def set_host(self, host):
        self.host = host or get_config("web.domain")

    @cached_property
    def file_info(self):
        return {}

    def file_url(self, fid, mid, name):
        if not self.file_info.get(fid):
            return ""
        url = Url(fid, mid, self.file_info[fid].qid, self.file_info[fid].tree_id, name)
        return f"- [file {fid}]({url})\r\n"

    def export_file(self, mold: int) -> None:
        report_dir = Path(tmp_storage.mount("error_reports"))
        if report_dir.exists():
            shutil.rmtree(report_dir)
        report_dir.mkdir(parents=True)

        for name, recall_error_fids in self.recall_errors.items():
            recall_error_fids = sorted(recall_error_fids)
            filepath = report_dir / f"{re.sub(r'/', '', name)}.md"
            with open(filepath, "w") as dumpfp:
                dumpfp.write(f"# {name}\r\n\n")
                dumpfp.write("答案错误\n")
                for fid in recall_error_fids:
                    dumpfp.write(self.file_url(fid, mold, name))

        for name, prec_error_fids in self.precision_errors.items():
            prec_error_fids = sorted(prec_error_fids)
            filepath = report_dir / f"{re.sub(r'/', '', name)}.md"
            with open(filepath, "a+") as dumpfp:
                dumpfp.write("\n预测答案比标准答案多\n")
                for fid in prec_error_fids:
                    dumpfp.write(self.file_url(fid, mold, name))

        for name, fit_fids in self.fit_stats.items():
            fit_fids = sorted(fit_fids)
            filepath = report_dir / f"{re.sub(r'/', '', name)}.md"
            with open(filepath, "a+") as dumpfp:
                dumpfp.write("\n预测正确\n")
                for fid in fit_fids:
                    dumpfp.write(self.file_url(fid, mold, name))

    async def export_stat_data(self, stat_result, mold, acid, fids, vid, diff_vid):
        aid_stat_result = {x["name"]: x for x in stat_result["result"]}
        total_percent = self.precision_format(stat_result["total_percent"])  # 叫准确率,但实际展示的是召回率
        export_dir = localstorage.mount(f"export_stat_result/{acid}")
        if not localstorage.exists(export_dir):
            localstorage.create_dir(export_dir)
        schema = Schema(mold.data)
        excel_path = f"{export_dir}/{mold.name}_测试报告_{datetime.now().strftime('%Y%m%d_%H:%M:%S')}.xlsx"
        workbook = openpyxl.Workbook()
        headers = [
            "文档ID",
            "文档名称",
            "一级字段",
            "准确率",
            "二级字段序号",
            "二级字段",
            "测试结果",
            "标注答案",
            f"预测答案\n(准确率:{total_percent})",
        ]
        sheet_data = []
        merge_range = defaultdict(list)
        current_row = 2
        combine_type = defaultdict(list)  # 组合类型（有二级字段）
        simple_type = {}

        for aid, items in self.export_data.items():
            if not items:
                continue
            answer = items[0].get("std") or items[0].get("preset_item")
            key_path = [i.split(":") for i in json.loads(answer["key"])[1:]]
            if len(key_path) == 2:
                top_field = key_path[0]
                combine_type[top_field[0]].extend(items)
            else:
                simple_type[aid] = items

        regroup_combine_type = self.regroup_data(combine_type)
        files = await pw_db.execute(NewFile.select(NewFile.id, NewFile.name).where(NewFile.id.in_(list(set(fids)))))
        file_names = {x.id: x.name for x in files}

        for field, sub in schema.iter_hierarchy():
            aid = field[-1]
            if not sub:
                items = simple_type.get(aid, [])
                current_row, merge_range = self.add_simple_data(
                    sheet_data, aid, items, aid_stat_result, current_row, file_names, merge_range, fids
                )
            else:
                aid_data = regroup_combine_type.get(aid)

                current_row, merge_range = self.add_combine_data(
                    sheet_data, aid, aid_data, aid_stat_result, current_row, file_names, merge_range, fids, sub
                )

        worksheet = dump_data_to_worksheet(workbook, headers, sheet_data)
        for col_name, value in merge_range.items():
            col = headers.index(col_name) + 1
            for start_row, end_row in value:
                worksheet.merge_cells(start_row=start_row, end_row=end_row, start_column=col, end_column=col)
        width_maps = {
            "A": 10,
            "B": 40,
            "C": 40,
            "D": 50,
            "E": 7,
            "F": 25,
            "G": 10,
            "H": 50,
            "I": 50,
        }
        if diff_vid:
            await self.adjust_xls_file(workbook, vid, diff_vid)
        self.optimize_xls_style(workbook, width_maps)
        workbook.save(excel_path)
        return excel_path

    @staticmethod
    async def adjust_xls_file(workbook, vid, diff_vid):
        model_version = await NewModelVersion.get_by_id(vid)
        diff_model_version = await NewModelVersion.get_by_id(diff_vid, include_deleted=True)
        if not (model_version and diff_model_version):
            raise CustomError("Invalid vid or diff_vid")
        worksheet = workbook.active
        worksheet.cell(1, 8).value = diff_model_version.name
        worksheet.cell(1, 9).value = model_version.name
        worksheet.column_dimensions["D"].hidden = True

    def get_empty_answer(self, fid):
        return {"fid": fid, "is_fit": False}

    @staticmethod
    def is_std_item(item):
        return item.get("std") and not item.get("preset_item")

    @staticmethod
    def is_preset_item(item):
        return item.get("preset_item") and not item.get("std")

    @staticmethod
    def is_fit(item, std_text, preset_text):
        if not item.get("std") and not item.get("preset_item"):
            return "一致"
        if not std_text and not preset_text:
            return "一致"
        return "一致" if item["is_fit"] else "不一致"

    def add_simple_data(self, sheet_data, aid, items, aid_stat_result, current_row, file_names, merge_range, fids):
        fid_count = len(fids)
        recall = self.get_recall(aid, aid_stat_result)
        merge_range["一级字段"].append((current_row, current_row + fid_count - 1))
        merge_range["准确率"].append((current_row, current_row + fid_count - 1))
        current_row += fid_count

        data_group_by_fid = defaultdict(list)
        for item in items:
            data_group_by_fid[item["fid"]].append(item)

        for fid in fids:
            fid_items = data_group_by_fid.get(fid, [self.get_empty_answer(fid)])
            if len(fid_items) == 2:
                std_item, preset_item = None, None
                for item in fid_items:
                    if self.is_std_item(item):
                        std_item = item
                    if self.is_preset_item(item):
                        preset_item = item

                if std_item and preset_item:  # 标注答案和预测答案都只有一条时,在excel里放到同一行
                    std_item["preset_item"] = preset_item["preset_item"]
                    fid_items = [std_item]

            for item in fid_items:
                std_text, preset_text = self.get_answer_text(item)
                sheet_data.append(
                    [
                        item["fid"],
                        file_names.get(item["fid"], "未找到对应文件"),
                        aid,
                        recall,
                        None,
                        None,
                        self.is_fit(item, std_text, preset_text),
                        std_text,
                        preset_text,
                    ]
                )

        return current_row, merge_range

    def add_combine_data(
        self, sheet_data, aid, aid_data, aid_stat_result, current_row, file_names, merge_range, fids, sub
    ):
        aid_data = aid_data or {}
        leaf_fields = [x[0][-1] for x in sub]
        leaf_count = len(leaf_fields)
        recall = self.get_recall(aid, aid_stat_result)
        aid_items_count = 0
        fid_row = current_row
        for fid in fids:
            empty_answer = []
            for leaf in leaf_fields:
                empty = self.get_empty_answer(fid)
                empty["second_field"] = leaf
                empty_answer.append(empty)

            empty_answer_map = {x["second_field"]: x for x in empty_answer}

            fid_data = aid_data.get(fid)
            if not fid_data:
                fid_data = {1: empty_answer}

            if len(fid_data) == 2:
                std_items, preset_items = None, None
                for items in fid_data.values():
                    if all(x.get("std") for x in items):
                        std_items = {x["second_field"]: x for x in items}
                    if all(self.is_preset_item(x) for x in items):
                        preset_items = {x["second_field"]: x for x in items}
                if std_items and preset_items:  # 标注答案和预测答案都只有一组时,在excel里放到同一行
                    duplicate_preset_items = []
                    for field, item in preset_items.items():
                        if field not in std_items:
                            std_items[field] = item
                        else:
                            if not std_items[field].get("preset_item"):
                                std_items[field]["preset_item"] = item["preset_item"]
                            else:
                                duplicate_preset_items.append(item)
                    fid_data = {1: std_items.values()}
                    if duplicate_preset_items:
                        fid_data[2] = duplicate_preset_items

            fid_items_count = 0
            index_row = fid_row
            for answer_index, items in fid_data.items():
                merge_range["二级字段序号"].append((index_row, index_row + leaf_count - 1))
                index_row += leaf_count
                aid_items_count += leaf_count
                fid_items_count += leaf_count
                items_map = {x["second_field"]: x for x in items}

                for leaf in leaf_fields:
                    item = items_map.get(leaf) or empty_answer_map[leaf]
                    std_text, preset_text = self.get_answer_text(item)
                    sheet_data.append(
                        [
                            item["fid"],
                            file_names.get(item["fid"], "未找到对应文件"),
                            aid,
                            recall,
                            answer_index,
                            item["second_field"],
                            self.is_fit(item, std_text, preset_text),
                            std_text,
                            preset_text,
                        ]
                    )

            merge_range["文档ID"].append((fid_row, fid_row + fid_items_count - 1))
            merge_range["文档名称"].append((fid_row, fid_row + fid_items_count - 1))

            fid_row += fid_items_count

        merge_range["一级字段"].append((current_row, current_row + aid_items_count - 1))
        merge_range["准确率"].append((current_row, current_row + aid_items_count - 1))
        current_row += aid_items_count

        return current_row, merge_range

    def get_recall(self, aid, aid_stat_result):
        if aid in aid_stat_result:
            recall = self.precision_format(aid_stat_result[aid].get("rate"))
            return recall

        data = []
        for key, value in aid_stat_result.items():
            if key.startswith(f"{aid}-"):
                recall = self.precision_format(value.get("rate"))
                data.append(f"{key.split('-')[-1]}: {recall}")
        if data:
            return "\n".join(data)

        return self.precision_format(0)

    def regroup_data(self, data):
        ret = defaultdict()
        for top_field, items in data.items():
            group_by_file = defaultdict(list)
            for item in items:
                group_by_file[item["fid"]].append(item)

            for fid, answers in group_by_file.items():
                group_by_file[fid] = self.group_by_index(answers)
            ret[top_field] = group_by_file

        return ret

    def group_by_index(self, items):
        std_items = defaultdict(list)
        preset_items = defaultdict(list)
        for item in items:
            if std := item.get("std"):
                key_path = [i.split(":") for i in json.loads(std["key"])[1:]]
                index = int(key_path[0][1]) + 1  # 导出结果里序号从1开始
                item["second_field"] = key_path[1][0]
                std_items[index].append(item)
            elif preset_item := item.get("preset_item"):
                key_path = [i.split(":") for i in json.loads(preset_item["key"])[1:]]
                index = int(key_path[0][1]) + 1  # 导出结果里序号从1开始
                item["second_field"] = key_path[1][0]
                preset_items[index].append(item)

            max_std_index = max(std_items.keys()) if std_items else 0
        for index, value in preset_items.items():
            std_items[max_std_index + index] = value

        return std_items

    def get_answer_text(self, item):
        std_text, preset_text = "", ""
        if std := item.get("std"):
            std_answer_item = AnswerItem(**std)
            std_text = std_answer_item.plain_text
        if label := item.get("preset_item"):
            label_answer_item = AnswerItem(**label)
            preset_text = label_answer_item.plain_text
        return std_text, preset_text

    def optimize_xls_style(self, workbook, width_maps):
        worksheet = workbook.active
        for col_index, width in width_maps.items():
            worksheet.column_dimensions[col_index].width = width
            for col in worksheet.iter_cols():
                for cell in col:
                    cell.font = Font(size=14)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def precision_format(precision):
        return "{:.0f}%".format((precision or 0) * 100)


error_report = ErrorReport()


def stat_label_correct(attr, label_data, strict, preset_item):
    label_correct = 0
    for l_data in label_data:
        box_correct = 0
        for p_data in preset_item["data"]:
            for l_box in l_data["boxes"]:
                if any(content_same(p_box["text"], l_box["text"], strict) for p_box in p_data["boxes"]):
                    box_correct += 1
                    continue
                if any(box_same(p_box["box"], l_box["box"], strict) for p_box in p_data["boxes"]):
                    box_correct += 1
                    continue
                if attr in FinancialAttribute.TARGET_ATTRS and any(
                    finanicial_tbline_same(p_box["text"], l_box["text"]) for p_box in p_data["boxes"]
                ):
                    box_correct += 1
                    continue
            if box_correct:
                label_correct += 1
                break
    return label_correct


def finanicial_tbline_same(text_a, text_b):
    text_a_vals = set(extract_number_vals(text_a))
    text_b_vals = set(extract_number_vals(text_b))
    return len(text_a_vals & text_b_vals) >= 2


def extract_number_vals(text):
    vals = []
    for _str in re.split(r"[\s]", text):
        val = try_parse_float(_str)
        if val is not None:
            vals.append(val)
    return vals


def try_parse_float(text):
    text = text.replace(",", "")
    multiplier = 1
    if "%" in text:
        text = text.replace("%", "")
        multiplier = 0.01
    try:
        val = float(text)
        return val * multiplier
    except ValueError:
        # print("%s is not a number" % text)
        return None


def box_same(box_a, box_b, strict):
    base = "element" if strict else "min"
    percent = PdfinsightReader.overlap_percent(box_a, box_b, base=base)
    # print(' box_same', attr, percent, percent >= 0.2)
    if percent >= (0.9 if strict else 0.2):
        return True
    return False


def content_same(text_a, text_b, strict):
    if isinstance(text_a, list):
        text_a = "".join(text_a)
    if isinstance(text_b, list):
        text_b = "".join(text_b)
    text_a = clean_txt(text_a)
    text_b = clean_txt(text_b)
    if not text_a or not text_b:
        return False
    if text_a == text_b:
        return True
    if not strict and (text_a in text_b or text_b in text_a):
        return True
    return False


def compare_preset_data(label_data, attr, predict, strict):
    if not predict:
        return False

    for preset_item in predict:
        if preset_item.get("mark"):
            continue
        label_correct = stat_label_correct(attr, label_data, strict, preset_item)
        if label_correct >= len(label_data):
            preset_item["mark"] = True
            return True

    return False


def valid_answer_item(answer_item):
    label_texts = []
    for data in answer_item["data"]:
        for box in data.get("boxes", []):
            text = box.get("text")
            if not text:
                continue
            label_texts.append(text)
    return bool(label_texts)


class AnswerMeasure:
    def __init__(self, answer_compare, answer_prepare):
        self.answer_compare = answer_compare
        self.answer_prepare = answer_prepare
        self.docs_count = 0
        self.fits = {}
        # 记录答案数量，用来计算 准确率
        self.attr_cnt = Counter()

    def append(self, fid, answer, standard, skip_reg):
        fits, attr_cnt = self.answer_compare(fid, answer, standard, skip_reg)
        self.attr_cnt.update(attr_cnt)
        for aid, val in fits.items():
            attr_fits = self.fits.setdefault(aid, {"name": aid, "fits": [], "idx_fits": []})
            attr_fits["fits"].extend(val["fits"])
            if val.get("idx_fits"):
                attr_fits["idx_fits"].extend(val["idx_fits"])

        self.docs_count += 1

    def measure_result(self):
        pass


class AnswerCompare:
    _white_list = []

    def __init__(self, strict=False, white_list_path=None):
        self.strict = strict
        self.white_list_path = white_list_path

    def __call__(self, fid, answer, standard, skip_reg=None):
        fits = {}
        attr_cnt = self.get_attr_cnt(answer, standard, skip_reg)
        errors_aid = []
        for aid, std_items in standard.items():
            errors_aid.append(aid)
            if self.is_skip_aid(aid, skip_reg):
                continue
            items = answer.get(aid, [])
            if len(items) > len(std_items):
                error_report.precision_errors[aid].add(fid)
            for std in std_items:
                if not any(x["boxes"] for x in std.get("data", [])):  # 没有boxes的答案无法参与统计
                    continue
                res = self.check(fid, aid, std, items)
                is_fit = res["fits"][0]
                error_report.export_data[aid].append(
                    {"fid": fid, "is_fit": is_fit, "std": std, "preset_item": res.get("preset_item", [])}
                )

                if is_fit is False:
                    error_report.recall_errors[aid].add(fid)
                else:
                    error_report.fit_stats[aid].add(fid)
                attr_fits = fits.setdefault(aid, {"name": aid, "fits": [], "idx_fits": []})
                attr_fits["fits"].append(res["fits"])
                if res.get("idx_fits"):
                    attr_fits["idx_fits"].append(res["idx_fits"])

            for item in items:
                if not item.get("mark"):  # 未能跟标注答案匹配上的
                    error_report.export_data[aid].append(
                        {"fid": fid, "is_fit": False, "std": None, "preset_item": item}
                    )

        # 没有标注答案的字段， 预测答案中该字段有值，属于多预测出答案的一种
        for aid, items in answer.items():
            if not items or all(not (item.get("data") or item.get("value")) for item in items):
                continue
            if aid not in errors_aid:
                error_report.precision_errors[aid].add(fid)
                for item in items:
                    error_report.export_data[aid].append(
                        {"fid": fid, "is_fit": False, "std": None, "preset_item": item}
                    )
        return fits, attr_cnt

    def get_attr_cnt(self, answer, standard, skip_reg=None):
        attr_cnt = Counter()
        for aid, items in answer.items():
            if self.is_skip_aid(aid, skip_reg):
                continue
            cnt = len([item for item in items if item and (item.get("data") or item.get("value"))])
            attr_cnt.update({aid: cnt})
        return attr_cnt

    def check(self, fid, aid, std, items):
        return {"fits": [], "idx_fits": None}

    def is_std_fit(self, attr, std, items):
        return False

    @staticmethod
    def re_aid_black_list():
        black_list = [re.compile(r"([>（]表格[<）])|(<.*>)")]
        return black_list

    @property
    def aid_black_list(self):
        return []

    def aid_in_black_list(self, aid):
        if aid in self.aid_black_list:
            return True
        re_black_list = self.re_aid_black_list()
        for black in re_black_list:
            if black.search(aid):
                return True

        return False

    @property
    def aid_white_list(self):
        if not self.white_list_path:
            return None

        if not self._white_list:
            with open(self.white_list_path) as fhp:
                self._white_list = json.load(fhp)
        return self._white_list

    def aid_in_white_list(self, aid):
        if self.aid_white_list:
            if not any(aid.startswith(x) for x in self.aid_white_list):
                return False

        return True

    def is_skip_aid(self, aid, skip_reg=None):
        if self.aid_in_black_list(aid):
            return True

        if not self.aid_in_white_list(aid):
            return True

        if skip_reg:
            if PatternCollection([skip_reg]).nexts(aid):
                return True

        return False


class PresetAnswerCompare(AnswerCompare):
    def re_aid_black_list(self):
        re_black_list = super(PresetAnswerCompare, self).re_aid_black_list()
        return re_black_list

    @property
    def aid_black_list(self):
        black = [
            "扉页-发行概况-股东公开发售股份数量",
            "扉页-发行概况-拟发行新股股票数量",
            "业务与技术-专利-权利人",
            "业务与技术-专利-专利授权国家",
            "业务与技术-专利-授权公告日",
            "业务与技术-专利-申请日",
            "业务与技术-专利-专利期限",
            "业务与技术-科研成果-项目名称",
            "业务与技术-科研成果-项目类型",
            "业务与技术-科研成果-角色",
            "本次发行概况-发行概况（三）-发行费用概算",
            "本次发行概况-发行概况（三）-超额配售发行上限",
            "发行人基本情况-间接控股股东-间接控股股东名称",
            "发行人基本情况-间接控股股东-间接控股股东类型",
        ]
        black.extend(super(PresetAnswerCompare, self).aid_black_list)
        return black

    def check(self, fid, aid, std, items):
        attr = aid.split("-")[-1]
        fit, preset_item = self.is_std_fit(attr, std, items)
        return {"fits": [fit], "preset_item": preset_item}

    def is_std_fit(self, attr, std, items):
        """
        std能否与items中某一条匹配
        :param attr:
        :param std:
        :param items:
        :return:
        """
        is_fit = False
        fit_item = None

        for item in items:
            if item.get("mark"):
                continue
            if self.is_fit(attr, std, item):
                item["mark"] = True
                is_fit = True
                fit_item = item
                break

        return is_fit, fit_item

    def is_fit(self, attr, std, item):
        count = 0
        for std_data in std["data"]:
            for data in item["data"]:
                if self.is_data_fit(attr, std_data, data):
                    count += 1
                    break

        return count == len(std["data"])

    def is_data_fit(self, attr, std_data, data):
        for std_box in std_data["boxes"]:
            if any(self.is_box_fit(attr, std_box, box, self.strict) for box in data["boxes"]):
                return True

        return False

    @staticmethod
    def is_box_fit(attr, std_box, box, strict):
        if content_same(std_box["text"], box["text"], strict):
            return True
        if box_same(std_box["box"], box["box"], strict):
            return True
        if attr in FinancialAttribute.TARGET_ATTRS:
            if finanicial_tbline_same(std_box["text"], box["text"]):
                return True

        return False


class ComplianceAnswerCompare(AnswerCompare):
    def check(self, fid, aid, std, items):
        attr = aid.split("-")[-1]
        fit = self.is_std_fit(attr, std, items)
        return {"fits": [fit]}

    def is_std_fit(self, attr, std, items):
        return std["value"] == items[0]["value"]


class CrudeAnswerCompare(AnswerCompare):
    overlap_threshold = 0.15

    def __init__(self, strict=False, white_list_path=None, **kwargs):
        super(CrudeAnswerCompare, self).__init__(strict, white_list_path)
        self.headnum = kwargs["headnum"]

    def get_attr_cnt(self, answer, standard, skip_reg=None):
        attr_cnt = Counter()
        for aid, items in answer.items():
            if self.is_skip_aid(aid, skip_reg):
                continue
            std_items = standard.get(aid, [])
            cnt = len([item for item in items if item])
            if std_items:
                cnt = cnt * len(std_items)
            attr_cnt.update({aid: cnt})
        return attr_cnt

    def get_headnum(self, aid):
        if any(x in aid for x in ["任职关系", "上下游关系", "交易对方所属区域"]):
            return 20
        return self.headnum

    def check(self, fid, aid, std, items):
        fits = []
        idx_fits = []
        for data in std.get("data", []):
            if not data or not data.get("boxes"):
                continue
            try:
                _idx_fits = self.compare_crude(aid, data, items)
            except ValueError as ex:
                logging.error("fid: %s, attr: %s, %s", fid, aid, ex)
                continue

            if _idx_fits:
                fits.append(True in _idx_fits)
                idx_fits.append(_idx_fits)
        return {"fits": fits, "idx_fits": idx_fits}

    def compare_crude(self, aid, data, crude_items):
        fits = []
        crude_nums = len(crude_items)
        for idx in range(self.get_headnum(aid)):
            fit = False
            if idx + 1 < crude_nums:
                crude = crude_items[idx]
                for box in data["boxes"]:
                    if int(box["page"]) != int(crude["page"]):
                        continue
                    if None in box["box"].values():
                        raise ValueError("box outline is None")
                    overlap = PdfinsightReader.overlap_percent(box["box"], crude["outline"], base="min")
                    if overlap > self.overlap_threshold:
                        fit = True
                        break
            fits.append(fit)
        return fits


class AnswerPrepare:
    @staticmethod
    def group_by_aid(answer):
        ret = defaultdict(list)
        for item in answer.get("items", []):
            aid = attribute_id(json.loads(item["key"]))
            ret[aid].append(item)
        return ret


class PresetAnswerPrepare(AnswerPrepare):
    async def __call__(self, question, answer, special_answer=None, old_version_answer=None):
        standard_answer, preset_answer = None, None

        if special_answer:
            preset_answer = self.group_by_aid(special_answer.data.get("userAnswer", {}))
        elif answer:
            preset_answer = self.group_by_aid(answer.get("userAnswer", {}))
        elif question.preset_answer:
            preset_answer = self.group_by_aid(question.preset_answer.get("userAnswer", {}))

        if old_version_answer:
            standard_answer = self.group_by_aid(old_version_answer.data.get("userAnswer", {}))
        elif answer_instance := await question.get_user_merged_answer():
            standard_answer = self.group_by_aid(answer_instance.get("userAnswer", {}))
        return preset_answer, standard_answer


class SSESchema5AnswerPrepare(AnswerPrepare):
    async def __call__(self, question, answer, *args):
        export_answer = await NewSpecialAnswer.get_answers(question.id, NewSpecialAnswer.ANSWER_TYPE_EXPORT)
        standard = export_answer[0].data if export_answer else None
        if not answer:
            predict_answer = await NewSpecialAnswer.get_answers(question.id, NewSpecialAnswer.ANSWER_TYPE_PREDICT)
            answer = predict_answer[0].data if predict_answer else None

        answer = self.group_by_aid(answer.get("userAnswer", {}))
        standard = self.group_by_aid(standard.get("userAnswer", {}))
        return answer, standard


class ComplianceAnswerPrepare(AnswerPrepare):
    async def __call__(self, question, answer, *args):
        standard = question.answer
        if not answer:
            answer = question.preset_answer

        answer = self.group_by_aid(answer.get("rule_result", {}))
        standard = self.group_by_aid(standard.get("rule_result", {}))
        return answer, standard


class CrudeAnswerPrepare(AnswerPrepare):
    def __init__(self, vid, threshold, headnum, **kwargs):
        super(CrudeAnswerPrepare, self).__init__(**kwargs)
        self.vid = vid
        self.threshold = threshold
        self.headnum = headnum

    def get_headnum(self, aid):
        if any(x in aid for x in ["任职关系", "上下游关系", "交易对方所属区域"]):
            return 20
        return self.headnum

    async def __call__(self, question, answer, *args):
        standard = question.answer
        standard = self.group_by_aid(standard.get("userAnswer", {}))

        crude_answer = {}

        if not answer:
            answer = question.crude_answer

        for aid in standard:
            headnum = self.get_headnum(aid)
            crude_answer[aid] = [c for c in answer.get(aid, []) if c["score"] > self.threshold][:headnum]

        return crude_answer, standard
