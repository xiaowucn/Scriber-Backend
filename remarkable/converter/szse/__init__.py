import copy
import json
import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from remarkable.answer.common import get_mold_name
from remarkable.config import project_root
from remarkable.converter import ConverterMaker, NullConverter
from remarkable.converter.szse.cyb_conv import CYBProspectusConverter
from remarkable.converter.utils import DataPack, ExcelWriter


class SZSEExcelWriter(ExcelWriter):
    p_col_name_tail = re.compile(r"（(包括|法人填写|法人不填|选填|必填|或有).*")
    temp_config = {
        "项目基本情况表": {
            "发行人信息": [
                {"mode": "table_kv", "answer_index": 0, "col_range": [1, 6], "row_range": [2, 21]},
            ],
            "发行前股本结构（万股）": [
                {
                    "mode": "table_kv",
                    "answer_index": 0,
                    "col_range": [2, 5],
                    "row_range": [22, 28],
                },
            ],
            "持股5%以上（含5%）股东信息": [
                {
                    "mode": "table_row",
                    "row_range": [29, -1],
                    "col_range": [3, 6],
                },
            ],
        },
        "联系方式": {
            "发行人": [
                {
                    "mode": "table_kv",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [2, 4],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [4, 9],
                },
            ],
            "保荐机构": [
                {
                    "mode": "table_kv",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [10, 11],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [11, 16],
                },
                {"mode": "table_kv", "answer_index": 1, "col_range": [1, 8], "row_range": [16, 17]},
                {
                    "mode": "table_tuple",
                    "answer_index": 1,
                    "col_range": [1, 8],
                    "row_range": [17, 22],
                },
                {
                    "mode": "table_kv",
                    "answer_index": 2,
                    "col_range": [1, 8],
                    "row_range": [22, 23],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 2,
                    "col_range": [1, 8],
                    "row_range": [23, 28],
                },
            ],
            "会计师事务所": [
                {
                    "mode": "table_kv",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [29, 30],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [30, 35],
                },
            ],
            "律师事务所": [
                {
                    "mode": "table_kv",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [36, 37],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [37, 42],
                },
            ],
            "资产评估机构": [
                {
                    "mode": "table_kv",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [43, 44],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 0,
                    "col_range": [1, 8],
                    "row_range": [44, 49],
                },
                {
                    "mode": "table_kv",
                    "answer_index": 1,
                    "col_range": [1, 8],
                    "row_range": [49, 50],
                },
                {
                    "mode": "table_tuple",
                    "answer_index": 1,
                    "col_range": [1, 8],
                    "row_range": [50, 55],
                },
            ],
        },
        "发行人相关人员情况": {
            "发行人相关人员情况": [
                {
                    "mode": "table_row",
                    "col_range": [1, 5],
                    "row_range": [1, -1],
                },
            ]
        },
    }

    @staticmethod
    def read_xls_template():
        xlsm_path = Path(project_root) / "data" / "szse_output" / "IPO企业基本信息与联系方式表.xlsm"
        workbook = load_workbook(xlsm_path, keep_vba=True)
        return workbook

    @classmethod
    def write(cls, data):
        json_answer = data["json_answer"]
        wb_template = cls.read_xls_template()
        financial_basic_data = json_answer["财务基础数据"][0]
        cls.write_financial_sheet(wb_template, financial_basic_data)

        # 第一节阶段上线要求先不写入这几个tab
        # for name, config in cls.temp_config.items():
        #     sheet = wb_template.get_sheet_by_name(name)
        #     sheet_data = json_answer[name][0]
        #     cls.write_non_financial_sheet(config, sheet, sheet_data)

        return wb_template

    @classmethod
    def write_non_financial_sheet(cls, config, sheet, sheet_data):
        for name, layouts in config.items():
            data = sheet_data.get(name)
            if not data:
                continue

            for idx, ans in enumerate(copy.deepcopy(data)):
                for key, value in ans.items():
                    data[idx][cls.clean_excel_col_name(key)] = value

            for layout in layouts:
                if layout["mode"] == "table_row":
                    cls.write_table_row(sheet, layout, data)
                elif layout["mode"] == "table_kv":
                    cls.write_table_kv(sheet, layout, data)
                elif layout["mode"] == "table_tuple":
                    cls.write_table_tuple(sheet, layout, data)

    @classmethod
    def write_financial_sheet(cls, wb_template, data):
        financial_basic_sheet = wb_template.get_sheet_by_name("财务基础数据")
        col_end = 5
        temp_config = {
            "合并资产负债表主要数据（万元）": {"row_range": [3, 12]},
            "合并利润表主要数据（万元）": {"row_range": [14, 27]},
            "合并现金流量表主要数据（万元）": {"row_range": [29, 34]},
            "最近三年一期主要财务指标表": {"row_range": [36, 53]},
            "其他指标": {"row_range": [55, 58]},
        }
        for name, conf in temp_config.items():
            item = data.get(name)
            if not item:
                # 其他指标被隐藏了
                continue
            for idx, ans in enumerate(item[::-1]):
                col = col_end - idx
                for row in range(*conf["row_range"]):
                    cell = financial_basic_sheet.cell(row, col)
                    # NOTE: we can't exec vba script to write the protected cell，so write it manually.
                    # if cell.protection.locked:  # 写保护
                    #     continue
                    key = financial_basic_sheet.cell(row, 1).value
                    key = key.replace("*", "")
                    value = ans.get(key)
                    if not value or not value.get("text"):
                        continue
                    cell.value = value["text"]

    @classmethod
    def clean_excel_col_name(cls, name):
        if not name:
            return name
        name = name.replace("*", "")
        name = cls.p_col_name_tail.sub("", name)
        return name


class SZSEConverterMaker(ConverterMaker):
    converter_map = {
        # schema_name: converter
        "招股说明书信息抽取": CYBProspectusConverter,
        "深交所信息抽取-创业板-注册制-财务基础数据": CYBProspectusConverter,
    }

    @classmethod
    def init(cls, answer):
        """
        加载转换类顺序:
            1. cls.converter_map[schema name]
            4. NullConverter(没有找到, 即暂未实现转换的 schema)
        """
        mold_name = get_mold_name(answer)
        if mold_name in cls.converter_map:
            return cls.converter_map[mold_name](answer)
        return NullConverter(answer)
