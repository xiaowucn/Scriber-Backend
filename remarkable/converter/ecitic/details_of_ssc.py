from collections import defaultdict

from openpyxl import Workbook

from remarkable.answer.node import AnswerItem
from remarkable.converter.ecitic.util import (
    Converter,
    TableItem,
    revise_multi_kv_items,
)


class DetailsOfSSC(Converter):
    SHEET_TITLE = "股东股份变更明细"
    sections = (SHEET_TITLE,)

    @classmethod
    def load_workbook_from_json(cls, sections) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = cls.SHEET_TITLE
        row_idx = 1
        for col_idx, header in enumerate(sections["headers"], 1):
            worksheet.cell(row_idx, col_idx, header)
        for page in sections["pages"].values():
            for row in page["rows"]:
                row_idx += 1
                for col_idx, cell in enumerate(row, 1):
                    # NOTE: None cell 是从哪来的？
                    worksheet.cell(row_idx, col_idx, AnswerItem(**(cell or {})).plain_text)
        return workbook

    def load_workbook(self) -> Workbook | TableItem:
        table_item = TableItem(
            "",
            None,
            data_handler=revise_multi_kv_items,
            sub_items=[
                TableItem(col, ["清单", col])
                for col in (
                    "股东名称",
                    "变更日期",
                    "证券代码",
                    "证券简称",
                    "托管单元代码",
                    "托管单元名称",
                    "证券类别",
                    "流通类型",
                    "权益类别",
                    "挂牌年份",
                    "变更股数",
                    "结余股数",
                    "变更摘要",
                )
            ],
        )

        table_item.rows = table_item.data_handler(self.answer_node, table_item)[1:]
        return table_item

    def to_tables(self):
        """
        {
            "headers": ['股东名称', '变更日期', '证券代码', '证券简称', '托管单元代码', '托管单元名称', '证券类别', '流通类型', '权益类别', '挂牌年份', '变更股数', '结余股数', '变更摘要'],
            "pages": {
                "0": {
                    "fill_status": 0,
                    "rows": [
                        [
                            {'box': [93.6338, 149.57, 103.0038, 120.47], 'text': '洪英亮'},
                            {'box': [103.0038, 149.57, 112.3738, 120.47], 'text': '2019-12-31'},
                            {'box': None, 'text': ''},
                            ...
                        ],
                        ...
                    ]
                },
                ...
            }
        }
        """
        assert isinstance(self.workbook, TableItem)
        headers, *rows = self.workbook.rows
        data = {"headers": headers, "pages": defaultdict(lambda: {"fill_status": 0, "rows": []})}
        for row in rows:
            line = defaultdict(list)
            for cell in row:
                if not cell.data or not cell.data[0].get("boxes"):
                    if line:
                        page = list(line)[-1]
                    elif data["pages"]:
                        page = list(data["pages"])[-1]
                    else:
                        page = "0"
                else:
                    page = str(cell.data[0]["boxes"][0]["page"])
                line[page].append(cell.to_dict())
            for page, lines in line.items():
                data["pages"][page]["rows"].append(lines)
        return data
