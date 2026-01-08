import os.path
from copy import copy
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_FORMULA, Cell
from openpyxl.worksheet.worksheet import Worksheet

from remarkable.config import project_root
from remarkable.converter.ecitic.util import (
    SECTION_LINE_COMMENT,
    Converter,
    CustomCell,
    TableItem,
    apply_formula,
    make_comment,
    nested_table_handler,
    revise_group_items,
    revise_multi_kv_items,
    revise_nested_items,
    revise_one_prefer_items,
    section_title_handler,
)


def new_important_debtor(worksheet: Worksheet, table: TableItem, row_idx: int):
    """新增重要债务人情况
    NOTE: 原文提不到, 纯拼接Excel内容
    """
    row_idx += 1
    worksheet.cell(row_idx, 1).value = table.title
    worksheet.cell(row_idx, 2).value = "名称"
    worksheet.cell(row_idx, 4).value = ""
    worksheet.cell(row_idx, 5).value = ""

    for key in ("基础资产现金流占比（%)", "经营情况"):
        row_idx += 1
        worksheet.cell(row_idx, 1).value = table.title
        worksheet.cell(row_idx, 2).value = key
        worksheet.cell(row_idx, 4).value = ""

    row_idx += 1
    worksheet.cell(row_idx, 1).value = table.title
    worksheet.cell(row_idx, 2).value = "主要财务情况"
    worksheet.cell(row_idx, 3).value = "科目"
    worksheet.cell(row_idx, 4).value = "报告期末情况"

    for key in ("净资产（万元）", "总负债（万元）", "总资产（万元）", "流动比率（%)", "速动比率（%)"):
        row_idx += 1
        worksheet.cell(row_idx, 1).value = table.title
        worksheet.cell(row_idx, 2).value = "主要财务情况"
        worksheet.cell(row_idx, 3).value = key
        worksheet.cell(row_idx, 4).value = ""

    for key in ("偿付能力", "资信水平"):
        row_idx += 1
        worksheet.cell(row_idx, 1).value = table.title
        worksheet.cell(row_idx, 2).value = key
        worksheet.cell(row_idx, 4).value = ""


class NotSmallScatteredConverter(Converter):
    sections = ("基本信息", "新增基础资产情况", "购买完成后基础资产情况", "其他信息")
    template_path = os.path.join(project_root, "data/ecitic/not_small_scattered_template.xlsx")
    sheet_title = "非“小而分散”类资产"

    @classmethod
    def load_workbook_from_json(cls, sections: list[list[list | Any]]) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = cls.sheet_title
        row_idx = 1
        for sec_title, tables in sections:
            if sec_title.get("name") not in (cls.sections[0], cls.sections[-1]):
                worksheet.cell(row_idx, 1).value = sec_title["name"]
                row_idx += 1
            for key, value in tables:
                if isinstance(key, str):
                    worksheet.cell(row_idx, 1).value = key
                    if "若有多个" in key:
                        row_idx += 1
                        worksheet.cell(row_idx, 1).value = value["text"]
                        worksheet.cell(row_idx, 1).comment = make_comment(value)
                    else:
                        worksheet.cell(row_idx, 2).value = value["text"]
                        worksheet.cell(row_idx, 2).comment = make_comment(value)
                    row_idx += 1
                elif isinstance(key, dict):
                    if key["mode"] == "new_debtor":  # 特殊表: 新增重要债务人情况
                        for t_key, t_value in value:
                            if isinstance(t_value, dict):
                                worksheet.cell(row_idx, 1).value = key["name"]
                                worksheet.cell(row_idx, 2).value = t_key
                                worksheet.cell(row_idx, 4).value = t_value["text"]
                                row_idx += 1
                            else:
                                for _key, _value in t_value:
                                    worksheet.cell(row_idx, 1).value = key["name"]
                                    worksheet.cell(row_idx, 2).value = t_key
                                    worksheet.cell(row_idx, 3).value = _key
                                    worksheet.cell(row_idx, 4).value = _value["text"]
                                    worksheet.cell(row_idx, 5).value = ""
                                    row_idx += 1
                    elif key["mode"] == "vertical":
                        worksheet.cell(row_idx, 1).value = key["name"]
                        row_idx += 1
                        for t_key, t_value in value[0]:  # 特殊表: 资产池基本情况
                            worksheet.cell(row_idx, 1).value = t_key
                            worksheet.cell(row_idx, 2).value = t_value["text"]
                            worksheet.cell(row_idx, 2).comment = make_comment(t_value)
                            row_idx += 1
                    else:
                        worksheet.cell(row_idx, 1).value = key["name"]
                        row_idx += 1
                        for idx, cols in enumerate(value):
                            if idx == 0:
                                for col_idx, col in enumerate((k[0] for k in cols), 1):
                                    worksheet.cell(row_idx, col_idx).value = col
                            row_idx += 1
                            for col_idx, col in enumerate((k[1] for k in cols), 1):
                                worksheet.cell(row_idx, col_idx).value = col["text"]
                                worksheet.cell(row_idx, col_idx).comment = make_comment(col)
                        row_idx += 1
        cls.apply_style(load_workbook(cls.template_path).active, worksheet)
        return workbook

    def load_workbook(self) -> Workbook:
        sections = [
            # section 1: 基本信息
            TableItem("报告名称", ["基本情况", "报告名称"]),
            TableItem("基础资产类型", None, ""),
            TableItem("基础资产是否为“小而分散”类资产", None, "否"),
            TableItem("循环购买频率", None, ""),
            TableItem("循环购买报告频率", None, ""),
            TableItem("循环购买时间", ["基本情况", "循环购买时间"]),
            TableItem(
                "可用于购买新增基础资产的价款总额（万元）", ["基本情况", "可用于购买新增基础资产的价款总额（万元）"]
            ),
            TableItem("可供购买的基础资产总额（万元）", ["基本情况", "可供购买的基础资产总额（万元）"]),
            TableItem("实际购买的新增基础资产总额（万元）", ["基本情况", "实际购买的新增基础资产总额（万元）"]),
            TableItem("基础资产买方（若有多个，以分号分割）", ["基本情况", "基础资产买方"]),
            TableItem("基础资产卖方（若有多个，以分号分割）", ["基本情况", "基础资产卖方"]),
            TableItem("循环购买账户资金划转情况", ["基本情况", "循环购买账户资金划转情况"]),
            # section 2: 新增基础资产情况
            TableItem("新增基础资产情况", None, None, xlsx_handler=section_title_handler),
            TableItem("新增基础资产数量（笔）", ["新增基础资产情况", "新增基础资产数量（笔）"]),
            TableItem("新增基础资产金额（万元）", ["新增基础资产情况", "新增基础资产金额（万元）"]),
            TableItem("新增基础资产债务人数量（个）", ["新增基础资产情况", "新增基础资产债务人数量（个）"]),
            TableItem(
                "新增债务人所在行业分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人所在行业", ["新增债务人所在行业分布", "债务人所在行业"]),
                    TableItem("未偿余额（万元）", ["新增债务人所在行业分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增债务人所在行业分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "新增债务人所属地区分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人所属地区", ["新增债务人所属地区分布", "债务人所属地区"]),
                    TableItem("未偿余额（万元）", ["新增债务人所属地区分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增债务人所属地区分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "新增重要债务人情况", None, data_handler=lambda x, y: list(range(11)), xlsx_handler=new_important_debtor
            ),
            # section 3: 购买完成后基础资产情况
            TableItem("购买完成后基础资产情况", None, None, xlsx_handler=section_title_handler),
            TableItem(
                "基础资产变化情况",
                None,
                data_handler=revise_nested_items,
                xlsx_handler=nested_table_handler,
                sub_items=[
                    TableItem(
                        "",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", None, "基础资产数量（笔）"),
                            TableItem("基础资产金额（万元）", None, "基础资产金额（万元）"),
                            TableItem("债务人数量（个）", None, "债务人数量（个）"),
                        ],
                    ),
                    TableItem(
                        "循环购买后",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", ["基础资产变化情况-循环购买后", "基础资产数量（笔）"]),
                            TableItem("基础资产金额（万元）", ["基础资产变化情况-循环购买后", "基础资产金额（万元）"]),
                            TableItem(
                                "债务人数量（个）",
                                None,
                                data_handler=revise_one_prefer_items,
                                sub_items=[
                                    TableItem("债务人数量（个）", ["基础资产变化情况-循环购买后", "债务人数量（个）"]),
                                    TableItem(
                                        "债务人数量（个）", ["基础资产变化情况-循环购买后", "基础资产数量（笔）"]
                                    ),
                                ],
                            ),
                        ],
                    ),
                    TableItem(
                        "报告期增减",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", None, ""),
                            TableItem("基础资产金额（万元）", None, ""),
                            TableItem("债务人数量（个）", None, ""),
                        ],
                    ),
                    TableItem(
                        "循环购买前",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", ["基础资产变化情况-循环购买前", "基础资产数量（笔）"]),
                            TableItem("基础资产金额（万元）", ["基础资产变化情况-循环购买前", "基础资产金额（万元）"]),
                            TableItem(
                                "债务人数量（个）",
                                None,
                                data_handler=revise_one_prefer_items,
                                sub_items=[
                                    TableItem("债务人数量（个）", ["基础资产变化情况-循环购买前", "债务人数量（个）"]),
                                    TableItem(
                                        "债务人数量（个）", ["基础资产变化情况-循环购买前", "基础资产数量（笔）"]
                                    ),
                                ],
                            ),
                        ],
                    ),
                ],
            ),
            TableItem(
                "资产池基本情况",
                None,
                data_handler=revise_group_items,
                sub_items=[
                    TableItem("项目", None, "数据情况"),
                    TableItem("未偿基础资产余额（万元）", ["基础资产变化情况-循环购买后", "基础资产金额（万元）"]),
                    TableItem("基础资产笔数（笔）", ["基础资产变化情况-循环购买后", "基础资产数量（笔）"]),
                    TableItem("债务人数量（户）", ["基础资产变化情况-循环购买后", "债务人数量（个）"]),
                    TableItem(
                        "单笔基础资产最高加权平均期限（月）", ["资产池基本情况", "单笔基础资产最高加权平均期限（月）"]
                    ),
                    TableItem("基础资产加权平均期限（月）", ["资产池基本情况", "基础资产加权平均期限（月）"]),
                    TableItem(
                        "单个债务人平均未偿基础资产余额（万元）",
                        ["资产池基本情况", "单个债务人平均未偿基础资产余额（万元）"],
                    ),
                    TableItem(
                        "单个债务人最高未偿基础资产余额（万元）",
                        ["资产池基本情况", "单个债务人最高未偿基础资产余额（万元）"],
                    ),
                    TableItem(
                        "前五大债务人未偿基础资产余额比例（%）",
                        ["资产池基本情况", "前五大债务人未偿基础资产余额比例（%）"],
                    ),
                    TableItem(
                        "前十大债务人未偿基础资产余额比例（%）",
                        ["资产池基本情况", "前十大债务人未偿基础资产余额比例（%）"],
                    ),
                ],
            ),
            TableItem(
                "债务人所在行业分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人所在行业", ["新增债务人所在行业分布", "债务人所在行业"]),
                    TableItem("未偿余额（万元）", ["新增债务人所在行业分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增债务人所在行业分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "债务人所属地区分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人所属地区", ["债务人所属地区分布", "债务人所属地区"]),
                    TableItem("未偿余额（万元）", ["债务人所属地区分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["债务人所属地区分布", "占比(%）"]),
                ],
            ),
            # section 4: 其他信息
            TableItem("其他与本报告事项相关且管理人认为应当披露的信息", None, ""),
        ]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_title
        row_idx = 0
        for section in sections:
            section.rows = section.data_handler(self.answer_node, section)
            section.xlsx_handler(worksheet, section, row_idx)
            row_idx += len(section.rows)

        self.apply_style(load_workbook(self.template_path).active, worksheet)

        return workbook

    @staticmethod
    def apply_style(from_ws, to_ws):
        to_ws_rows = dict(enumerate(to_ws.rows, 1))
        rows = []
        for row_idx in range(1, to_ws.max_row + 1):
            cell = to_ws.cell(row_idx, 1)
            if CustomCell.load_comment(cell) == SECTION_LINE_COMMENT:
                rows.append([to_ws_rows[row_idx]])
            elif cell.comment is not None and rows[-1][-1][0].comment is not None:
                rows[-1].append(to_ws_rows[row_idx])
            else:
                rows.append([to_ws_rows[row_idx]])

        rows_map = dict(enumerate(rows, 1))
        for idx, _rows in enumerate(rows, 1):
            for row in _rows:
                valid_cells = [c for c in row if isinstance(c, Cell) and c.value is not None]
                valid_cells_len = len(valid_cells)
                first_cell = valid_cells[0]
                if first_cell.value != "新增重要债务人情况" and valid_cells_len > 2:
                    # 二三列有合并, 从第三列起需要整体往右挪一列
                    for _idx in range(valid_cells_len + 1, 3, -1):
                        to_ws.cell(first_cell.row, _idx).value = to_ws.cell(first_cell.row, _idx - 1).value
                        to_ws.cell(first_cell.row, _idx).comment = to_ws.cell(first_cell.row, _idx - 1).comment
                    to_ws.cell(first_cell.row, 3).value = None

                for col_idx, cell in enumerate(row, 1):
                    from_ws_cell = from_ws.cell(idx, col_idx)
                    cell.font = copy(from_ws_cell.font)
                    cell.border = copy(from_ws_cell.border)
                    cell.fill = copy(from_ws_cell.fill)
                    cell.number_format = copy(from_ws_cell.number_format)
                    cell.protection = copy(from_ws_cell.protection)
                    cell.alignment = copy(from_ws_cell.alignment)
                    if from_ws_cell.data_type == TYPE_FORMULA and cell.comment is None:
                        apply_formula(rows_map[idx], cell, from_ws_cell.value)
                # 行高
                to_ws.row_dimensions[row[0].row].height = copy(from_ws.row_dimensions[idx].height)
        # 合并单元格
        for merged_cell in from_ws.merged_cells.ranges:
            row_idx = merged_cell.left[0][0]
            from_col_idx, to_col_idx = merged_cell.left[0][1], merged_cell.right[0][1]
            row_offset = merged_cell.bottom[0][0] - merged_cell.top[0][0]
            for row in rows_map[row_idx]:
                to_ws.merge_cells(
                    start_row=row[0].row,
                    start_column=from_col_idx,
                    end_row=row[0].row + row_offset,
                    end_column=to_col_idx,
                )
        for idx in ("A", "B", "C", "D", "E"):
            # 列宽
            to_ws.column_dimensions[idx].width = copy(from_ws.column_dimensions[idx].width)
