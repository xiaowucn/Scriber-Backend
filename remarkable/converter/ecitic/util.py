# CYC: skip-file
import decimal
import io
import json
import logging
import re
from collections import defaultdict
from copy import deepcopy
from enum import IntEnum
from functools import reduce
from itertools import zip_longest
from typing import Any, Callable

from attrs import Factory, define, field
from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.comments import Comment
from openpyxl.descriptors import Descriptor
from openpyxl.worksheet.worksheet import Worksheet

from remarkable.answer.node import AnswerItem, AnswerNode
from remarkable.common.constants import FillInStatus
from remarkable.common.pattern import PatternCollection
from remarkable.common.util import clean_txt, compact_dumps
from remarkable.converter import BaseConverter, DataPack
from remarkable.converter.utils import cn_unit
from remarkable.pw_models.model import NewSpecialAnswer
from remarkable.rule.rule import revise_answer

COMMENT_AUTHOR = "PAI_SCRIBER"
SECTION_LINE_COMMENT = "section_line_comment"
to_unit_pattern = re.compile(r"[（\(](.*?)[）)]$")
amount_unit_pattern = re.compile(r"(?P<val>-?\d+(,\d+)*(\.\d+)?)(?P<unit>\D*)$")
logger = logging.getLogger(__name__)


class CustomCell:
    @classmethod
    def obj2dict(cls, obj):
        ret = {}
        for key in obj.__attrs__ + obj.__elements__:
            value = getattr(obj, key)
            if isinstance(value, Descriptor):
                value = None
            if getattr(value, "__module__", "").startswith("openpyxl."):
                value = cls.obj2dict(value)
            ret[key] = value
        return ret

    @classmethod
    def cell2dict(cls, cell: Cell | MergedCell):
        """
        {
            "row": 1,
            "column": 1,
            "value": "报告名称",
            "coordinate": "A1",
            "font": {
                "name": "宋体",
                "b": true,
                "i": false,
            },
            "alignment": {
                "horizontal": null,
                "vertical": null,
                "textRotation": 0,
                "wrapText": null,
                "shrinkToFit": null,
                "indent": 0.0,
                "relativeIndent": 0.0,
                "justifyLastLine": null,
                "readingOrder": 0.0
            },
            "comment": {
                "data": [],
                "value": "",
                "key": "[\"“小而分散”类资产:0\",\"基本情况:0\",\"报告名称:0\"]",
                "schema": {},
                "marker": {}
            }
        }
        """
        ret = {}
        for key in ("row", "column", "value", "coordinate"):
            ret[key] = getattr(cell, key)
        for key in ("font", "alignment"):
            ret[key] = cls.obj2dict(getattr(cell, key))
        ret["comment"] = cls.load_comment(cell)
        return ret

    @classmethod
    def cell2dict_legacy(cls, cell: Cell | MergedCell):
        comment = cls.load_comment(cell)
        if not isinstance(comment, dict):
            return {"text": cell.value}
        ret = {
            "data": comment["data"],
            "text": cell.value,
            "value": comment["value"],
            "key": comment["key"],
            "schema_path": "|".join(i.split(":")[0] for i in json.loads(comment["key"])),
            "manual": comment["manual"],
        }
        for key in "manual", "meta":
            if comment.get(key) is not None:
                ret[key] = comment[key]
        return ret

    @staticmethod
    def load_comment(cell: Cell | MergedCell):
        return json.loads(cell.comment.text) if cell.comment and cell.comment.author == COMMENT_AUTHOR else None


def revise_multi(from_unit: str | None, to_unit: str | None):
    """获取单位换算倍率"""
    multi = 1
    for word in from_unit or "":
        multi *= cn_unit.get(word, 1)

    for word in to_unit or "":
        multi /= cn_unit.get(word, 1)
    return decimal.Decimal(multi).quantize(decimal.Decimal(f"{multi}"))


def guess_unit(dates: list[str], units: list[str]):
    if units and units[0]:
        return units[0]
    if max(float(d) for d in dates) > 12:
        # 无单位情况下：(0,X] X>12时，认为单位为天
        # https://gitpd.paodingai.com/cheftin/docs_scriber/-/issues/489#note_136185
        return "天"
    return "月"


def convert_date(data: AnswerItem) -> str:
    """日期描述中的单位转换
    1. 100天以内 -> 3月以内
    2. 30-90天 -> 1-3月
    """
    date_str, *units = (clean_txt(t.split(":")[-1].strip()) for t in data.plain_text.split("|"))
    if re.search(r"^\d+,\d+\.?\d+$", date_str):
        date_str = date_str.replace(",", "")
    dates = re.findall(r"\d+\.?\d*", date_str, 2)
    if not dates:
        return date_str

    unit = guess_unit(dates, units)
    if len(dates) == 1:
        dst_unit, dst_val = revise_amount(data, dates[0] + unit)
        if dst_unit:
            date_str = date_str.replace(unit, dst_unit)
        return re.sub(r"(.*?)\d+\.?\d*(.*?)", r"\1__from__\2", date_str, count=2).replace(
            "__from__", dst_val.to_eng_string()
        )
    from_str, to_str = dates
    dst_from_unit, dst_from_val = revise_amount(data, from_str + unit)
    dst_to_unit, dst_to_val = revise_amount(data, to_str + unit)
    dst_unit = dst_from_unit or dst_to_unit
    if dst_unit:
        date_str = date_str.replace(unit, dst_unit)
    return (
        re.sub(r"(.*?)\d+\.?\d*(.*?)\d+\.?\d*(.*?)", r"\1__from__\2__to__\3", date_str, count=2)
        .replace("__from__", dst_from_val.to_eng_string())
        .replace("__to__", dst_to_val.to_eng_string())
    )


def revise_amount(data_item: AnswerItem, val=None) -> tuple[str, decimal.Decimal]:
    keys = [key.split(":")[0] for key in json.loads(data_item.key)]
    # 币种: 人民币|数值: 10,000.00|单位: 万元 -> 10,000.00万元
    if val is not None:
        amount = val
    elif data_item.plain_text.startswith("币种"):
        amount = clean_txt("".join(t.split(":")[-1].strip() for t in data_item.plain_text.split("|"))[1:])
    else:
        amount = clean_txt("".join(t.split(":")[-1].strip() for t in data_item.plain_text.split("|")))

    # 取出目标字段所需单位
    to_unit = None
    matches = list(filter(None, (to_unit_pattern.search(k) for k in keys)))
    if matches:
        to_unit = matches[-1].group(1)

    # 转换剩余数值单位
    match = amount_unit_pattern.search(amount)
    if not match:
        return to_unit, decimal.Decimal("0")
    val = decimal.Decimal(match.group("val").replace(",", "") or "0")
    unit = match.group("unit") or to_unit
    # 有效数字如无特殊要求，均保持跟原文字段一致
    ret_val = val * revise_multi(unit, to_unit)
    ret = ret_val.quantize(
        decimal.Decimal("0.00") if data_item.plain_text.startswith("月份:") else val, rounding=decimal.ROUND_HALF_UP
    )
    # 120元 => 0.012万元 != 0万元 需要保留原始计算值，不做有效数字处理
    return to_unit, ret_val if ret == 0 and ret_val != 0 else ret


def is_amount_item(data_item: AnswerItem) -> bool:
    return (
        not data_item.is_empty and data_item.key and any(i in data_item.key for i in ("数值", "单位", "金额", "币种"))
    )


def is_date_item(data_item: AnswerItem) -> bool:
    return (
        not data_item.is_empty
        and data_item.key
        and any(
            i in data_item.key
            for i in (
                "（月）",
                "（日）",
                "（天）",
            )
        )
    )


def revise_value(data_item: str | AnswerItem) -> str:
    if isinstance(data_item, str):
        return data_item
    if is_date_item(data_item):
        return convert_date(data_item)
    if "占比" in data_item.key and "%" in data_item.plain_text:
        # 百分比不做单位转换
        return data_item.simple_text(enum=False).replace("%", "") + "%"
    if is_amount_item(data_item):
        return revise_amount(data_item)[-1].to_eng_string()
    return data_item.plain_text


def make_comment(data_item: str | dict | AnswerItem, force=False) -> Comment | None:
    if isinstance(data_item, AnswerItem):
        pass
    elif isinstance(data_item, dict):
        if "data" in data_item:
            data_item = AnswerItem(item=data_item)
        else:
            data_item = data_item["text"]
    comment = compact_dumps(data_item)
    if force:
        return Comment(comment, COMMENT_AUTHOR)
    if isinstance(data_item, str):
        return None
    return Comment(comment, COMMENT_AUTHOR)


def empty_answer_with_path(parent_node: AnswerNode, data_item: "TableItem") -> AnswerItem:
    path = parent_node.path + [(i, 0) for i in data_item.key_path]
    return AnswerItem(key=compact_dumps([f"{k}:{i}" for k, i in path]))


def section_title_handler(worksheet: Worksheet, table: "TableItem", row_idx: int):
    """分界行"""
    row_idx += 1
    cell = worksheet.cell(row_idx, 1)
    cell.value = table.title
    cell.comment = make_comment(SECTION_LINE_COMMENT, force=True)


def nested_table_handler(worksheet: Worksheet, table: "TableItem", row_idx: int):
    first_row, *others = table.rows
    cell = worksheet.cell(row_idx + 1, 1)
    cell.value = revise_value(first_row[0])
    cell.comment = make_comment(first_row[0])
    for idx, row in enumerate(others, 1):
        for col_idx, item in enumerate(row, 1):
            cell = worksheet.cell(row_idx + idx + 1, col_idx)
            cell.value = revise_value(item)
            cell.comment = make_comment(item)


def default_excel_handler(worksheet: Worksheet, table: "TableItem", row_idx: int):
    if table.data_handler.__name__ == "revise_multi_kv_items":
        first_row, *others = table.rows
        cell = worksheet.cell(row_idx + 1, 1)
        cell.value = revise_value(first_row[0])
        cell.comment = make_comment(first_row[0])
        for idx, row in enumerate(others, 1):
            for col_idx, item in enumerate(row, 1):
                cell = worksheet.cell(row_idx + idx + 1, col_idx)
                cell.value = revise_value(item)
                cell.comment = make_comment(item)
        return
    if table.data_handler.__name__ == "revise_group_items":
        first_row, *others = table.rows
        cell = worksheet.cell(row_idx + 1, 1)
        cell.value = revise_value(first_row[0])
        cell.comment = make_comment(first_row[0])

        for idx, (first_col, second_col) in enumerate(others, 1):
            cell = worksheet.cell(row_idx + idx + 1, 1)
            cell.value = revise_value(first_col)
            cell.comment = make_comment(first_col)

            cell = worksheet.cell(row_idx + idx + 1, 2)
            cell.value = revise_value(second_col)
            cell.comment = make_comment(second_col)
        return
    rows_len = len(table.rows)
    if rows_len == 2:
        # 两行一列的情况
        # 合并首行四列单元格
        cell = worksheet.cell(row_idx + 1, 1)
        cell.value = revise_value(table.rows[0][0])
        cell.comment = make_comment(table.rows[0][0])

        # value 在下一行, 同样合并四列单元格
        cell = worksheet.cell(row_idx + 2, 1)
        cell.value = (
            table.rows[1] if isinstance(table.rows[1], str) else "；".join(item.plain_text for item in table.rows[1])
        )
        cell.comment = make_comment(reduce(lambda x, y: x + y, table.rows[1]))
        return

    if rows_len == 1:
        # 一行两列的情况
        first_cell, second_cell = table.rows[0]
        cell = worksheet.cell(row=row_idx + 1, column=1)
        cell.value = revise_value(first_cell)
        cell.comment = make_comment(first_cell)

        cell = worksheet.cell(row=row_idx + 1, column=2)
        cell.value = revise_value(second_cell)
        cell.comment = make_comment(second_cell)
        return


def revise_answer_node(node: AnswerNode) -> AnswerItem:
    if not node.isleaf():
        return revise_answer({0: node}).data
    if node.name.endswith("（月）"):
        for key, value in node.parent.items():
            if key[0].endswith("-单位"):
                node.data.plain_text = f"月份: {node.data.plain_text}|单位: {value.data.plain_text}"
                break
    return node.data


def revise_kv_items(
    answer_node: AnswerNode, row_item: "TableItem", parent_node: AnswerNode | None = None
) -> list[list[str | AnswerItem]]:
    if row_item.key_path is None:
        # 不参与提取的字段，直接填充默认值
        if row_item.default_content is None:
            return [[row_item.title]]
        return [
            [
                row_item.title,
                DocType.which_value(row_item.doc_type, row_item.title) or row_item.default_content or "",
            ]
        ]

    if parent_node is None:
        parent_node = answer_node

    if parent_node is None:
        logger.error("Possible empty answer detected.")
        return []

    if len(row_item.key_path) > 1:
        copy_item = deepcopy(row_item)
        copy_item.key_path = row_item.key_path[1:]
        rows = []
        for node in parent_node.get(row_item.key_path[0], defaultdict(dict)).values():
            rows.extend(revise_kv_items(answer_node, copy_item, node))

        # KV分行输出的情况
        if rows and "若有多个" in row_item.title:
            return [
                [row_item.title],
                DocType.which_value(row_item.doc_type, row_item.title) or [r[-1] for r in rows],
            ]
        return rows if rows else [[row_item.title, empty_answer_with_path(parent_node, row_item)]]

    for node in parent_node.get(row_item.key_path[0], defaultdict(dict)).values():
        return [[row_item.title, revise_answer_node(node)]]
    return [[row_item.title, empty_answer_with_path(parent_node, row_item)]]


class DocType(IntEnum):
    def __new__(cls, value: int, patterns: tuple[str] | None, description=""):
        obj = int.__new__(cls, value)
        obj._value_ = value

        obj.pattern = PatternCollection(patterns)
        obj.description = description
        return obj

    NORMAL = (0, None, "普通")
    MICRO_FINANCE = (1, ("荟享", "耘睿", r"博远\d+号第[2二]期", r"[小微][额贷]", r"(360|三六[零〇O])数科"), "小微贷款")
    RECEIVABLE_ACCOUNTS = (2, ("东道", "禾昱", r"博远\d+号第[13一三]期"), "应收账款")

    @classmethod
    def which_type(cls, texts: str | list[str]) -> "DocType":
        if isinstance(texts, str):
            texts = [texts]
        if any(cls.MICRO_FINANCE.pattern.nexts(t) for t in texts):
            return cls.MICRO_FINANCE
        if any(cls.RECEIVABLE_ACCOUNTS.pattern.nexts(t) for t in texts):
            return cls.RECEIVABLE_ACCOUNTS
        return cls.NORMAL

    @classmethod
    def which_value(cls, doc_type: "DocType", title: str) -> str:
        # 需求依据：https://gitpd.paodingai.com/cheftin/docs_scriber/-/issues/484#note_159450
        # 变更依据：https://gitpd.paodingai.com/cheftin/docs_scriber/-/issues/2143#note_324911
        default_values = {
            cls.MICRO_FINANCE: {
                "基础资产类型": "债权类 小额贷款债权 互联网小额贷款债权",
                "基础资产买方（若有多个，以分号分割）": "中信证券股份有限公司（代本专项计划）",
            },
            cls.RECEIVABLE_ACCOUNTS: {
                "基础资产类型": "债权类 应收账款 一般企业应收账款",
            },
        }
        return default_values.get(doc_type, {}).get(title) or ""


@define
class TableItem:
    title: str = field()
    key_path: list[str] | None = field()
    default_content: str | None = field(default=None)
    rows: list[list[str | AnswerItem]] = field(default=Factory(list))
    data_handler: Callable | None = field(default=revise_kv_items)
    xlsx_handler: Callable | None = field(default=default_excel_handler)
    sub_items: list = field(default=Factory(list))
    doc_type: DocType = field(default=DocType.NORMAL)


class Converter(BaseConverter):
    def __init__(self, meta_data: DataPack, debug=False):
        super().__init__(meta_data.answer)
        self.debug = debug
        self.meta_data = meta_data
        self.workbook: Workbook | TableItem = self.load_workbook()

    async def work(self):
        await NewSpecialAnswer.update_or_create(
            self.meta_data.question.id, NewSpecialAnswer.ANSWER_TYPE_JSON, self.to_tables()
        )

    @classmethod
    def load_workbook_from_json(cls, sections: list[list[list | Any]]) -> Workbook:
        raise NotImplementedError

    def load_workbook(self) -> Workbook:
        raise NotImplementedError

    @staticmethod
    def clean_comment(workbook: Workbook):
        # Remove our special comments
        sheet = workbook.active
        for cell in (cell for row in sheet.rows for cell in row):
            if cell.comment and cell.comment.author == COMMENT_AUTHOR:
                cell.comment = None

    @classmethod
    def to_excel(cls, workbook: Workbook | None, keep_comment=False) -> bytes:
        if not keep_comment:
            cls.clean_comment(workbook)
        with io.BytesIO() as io_fp:
            workbook.save(io_fp)
            return io_fp.getvalue()

    def to_dict(self) -> dict:
        """TODO: 通用的表格数据处理流实现
        {
            "max_row": int,
            "max_column": int,
            "merged_cell_ranges": ["B1:D1", "B2:D2", ...],
            "cells": [[ROW1], [ROW2], ...]
        }
        """
        sheet = self.workbook.active
        ret = {k: getattr(sheet, k) for k in ("max_row", "max_column")}
        ret["merged_cell_ranges"] = [str(cell) for cell in sheet.merged_cell_ranges]
        ret["cells"] = []
        for row in sheet.rows:
            ret["cells"].append([CustomCell.cell2dict(cell) for cell in row])
        return ret

    def to_tables(self) -> list[list[list | Any]]:
        """
        [
            [
                "基本信息",
                [
                    [
                        "报告名称",
                        {
                            "data": [...],
                            "text": "关于万科-中信证券购房尾款资产支持专项计划第十九期第一次循环购买的公告",
                            "value": "",
                            "schema_path": "“小而分散”类资产|基本情况|报告名称",
                            "manual": true
                        }
                    ],
                    [
                        "基础资产类型",
                        {
                            "text": "债权类 保理融资债权",
                        }
                    ],

                ]
            ],
            [
                "新增基础资产情况",
                [...]
            ],
            [
                "购买完成后基础资产情况",
                [...]
            ],
            [
                "其他信息",
                [...]
            ],
        ]
        """
        worksheet = self.workbook.active
        section_key = self.sections[0]
        tables = {section_key: []}
        table_rows = dict(enumerate(worksheet.rows, 1))
        new_debtor_info = {}
        for idx, row in table_rows.items():
            if CustomCell.load_comment(row[0]) == SECTION_LINE_COMMENT or row[0].value in self.sections:
                section_key = row[0].value
                tables[section_key] = []
                continue
            valid_cells = [c for c in row if isinstance(c, Cell) and c.value is not None]
            next_line_cells = [c for c in table_rows.get(idx + 1, []) if isinstance(c, Cell)]
            if section_key != self.sections[0] and len(valid_cells) == 1 and len(next_line_cells) > 1:
                tables[section_key].append([valid_cells[0].value, {}])
                continue

            if valid_cells[0].value == "新增重要债务人情况":
                new_debtor_info[valid_cells[1].value] = [CustomCell.cell2dict_legacy(valid_cells[2])]
                continue
            if new_debtor_info:
                if valid_cells[0].value == "主要财务情况" or (
                    isinstance(row[1], MergedCell) and isinstance(row[2], Cell)
                ):
                    new_debtor_info.setdefault("主要财务情况", defaultdict(list))[valid_cells[-2].value].append(
                        CustomCell.cell2dict_legacy(valid_cells[-1])
                    )
                else:
                    new_debtor_info[valid_cells[0].value] = [CustomCell.cell2dict_legacy(valid_cells[1])]
                if next_line_cells[0].value == self.sections[2]:
                    tables[self.sections[1]].append(["新增重要债务人情况", new_debtor_info])
                    new_debtor_info = None
                continue

            if len(valid_cells) == 1:
                if len(tables[section_key][-1]) == 1:
                    tables[section_key][-1].append([CustomCell.cell2dict_legacy(valid_cells[0])])
                else:
                    tables[section_key].append([valid_cells[0].value])
            elif len(valid_cells) == 2:
                if valid_cells[0].value == "其他与本报告事项相关且管理人认为应当披露的信息":
                    # HARDCODE: fill in the missing section title
                    tables[self.sections[-1]] = [[valid_cells[0].value, [CustomCell.cell2dict_legacy(valid_cells[1])]]]
                elif tables[section_key] and isinstance(tables[section_key][-1][-1], dict):
                    tables[section_key][-1][-1][valid_cells[0].value] = [CustomCell.cell2dict_legacy(valid_cells[1])]
                else:
                    tables[section_key].append([valid_cells[0].value, [CustomCell.cell2dict_legacy(valid_cells[1])]])
            else:
                _table: dict = tables[section_key][-1][-1]
                if not _table:
                    _table.update({k.value: [] for k in valid_cells})
                else:
                    for cell, cell_list in zip(valid_cells, _table.values()):
                        cell_list.append(CustomCell.cell2dict_legacy(cell))

        return self.dict2list(tables)

    @staticmethod
    def dict2list(table):
        ret = []
        for section_key, pairs in table.items():
            ret.append([{"name": section_key, "fill_status": FillInStatus.TODO}, []])
            for key, values in pairs:
                if isinstance(values, list):
                    ret[-1][-1].append([key, values[0]])
                elif key == "新增重要债务人情况":
                    ret[-1][-1].append(
                        [
                            {
                                "name": key,
                                "mode": "new_debtor",
                                "allow_clone_row": False,
                            },
                            [
                                [
                                    k,
                                    vals[0]
                                    if isinstance(vals, list)
                                    else [list(zip(vals.keys(), v)) for v in zip(*vals.values())][0],
                                ]
                                for k, vals in values.items()
                            ],
                        ]
                    )
                else:
                    ret[-1][-1].append(
                        [
                            {
                                "name": key,
                                "mode": "vertical" if key == "资产池基本情况" else "horizontal",
                                "allow_clone_row": key not in ("资产池基本情况", "基础资产变化情况"),
                            },
                            [list(zip(values.keys(), v)) for v in zip(*values.values())],
                        ]
                    )
        return ret


def revise_multi_kv_items(answer_node: AnswerNode, row_item: TableItem) -> list[list[str | AnswerItem]]:
    rows = [[row_item.title], [i.title for i in row_item.sub_items]]
    for item in row_item.sub_items:
        item.rows = item.data_handler(answer_node, item)
    rows.extend(
        zip_longest(*((row[1] for row in sub_item.rows) for sub_item in row_item.sub_items), fillvalue=AnswerItem())
    )
    return rows


def revise_group_items(answer_node: AnswerNode, row_item: TableItem) -> list[list[str | AnswerItem]]:
    rows = [[row_item.title]]
    for item in row_item.sub_items:
        rows.extend(item.data_handler(answer_node, item))
    return rows


def revise_one_prefer_items(answer_node: AnswerNode, row_item: TableItem) -> list[list[str | AnswerItem]]:
    """Parse answer from multi key paths, return the first match"""
    rows = []
    for idx, item in enumerate(row_item.sub_items):
        if (
            idx > 0
            and row_item.doc_type == DocType.RECEIVABLE_ACCOUNTS
            and item.title
            in {
                "新增基础资产债务人数量（个）",
                "债务人数量（个）",
            }
        ):
            # 应收账款类，首选字段未取到值时，不再继续提取
            continue
        rows = item.data_handler(answer_node, item)
        if not rows or not rows[-1] or rows[-1][-1].is_empty:
            continue
        if "若有多个" not in item.title and idx > 0:
            for value in (row[1] for row in rows):
                value["meta"] = {
                    "tips": "该信息取自报告期、请确认是否准确"
                    if row_item.title == "循环购买时间"
                    else "文档未提及，该信息取自基础资产信息、请确认是否正确"
                }
        if rows:
            break
    return rows


def revise_nested_items(answer_node: AnswerNode, row_item: TableItem) -> list[list[str | AnswerItem]]:
    rows = [[row_item.title], [i.title for i in row_item.sub_items]]
    for item in row_item.sub_items:
        item.rows = item.data_handler(answer_node, item)
    rows.extend(zip_longest(*(sub_item.rows[2] for sub_item in row_item.sub_items), fillvalue=AnswerItem()))
    return rows


def subtraction(row: tuple[Cell], cell: Cell, formula: str):
    cell_a = cell_b = None
    # (B37,D37) -> ['B42', 'D42']
    for coordinate in re.sub(r"(\d+)", str(cell.row), formula.rstrip(")").lstrip("(")).split(","):
        for _cell in row:
            if _cell.coordinate != coordinate:
                continue
            if cell_a is None:
                cell_a = _cell
                continue
            if cell_b is None:
                cell_b = _cell
                break
    if cell_a.value and cell_b.value:
        cell.value = (decimal.Decimal(cell_a.value) - decimal.Decimal(cell_b.value)).to_eng_string()


def apply_formula(rows: list[tuple[Cell]], cell: Cell, formula: str):
    func_map = {
        "=IMSUB": subtraction,
    }
    for row in rows:
        for key, func in func_map.items():
            if formula.startswith(key):
                func(row, cell, formula.replace(key, ""))
