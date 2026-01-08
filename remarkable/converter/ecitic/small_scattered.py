import os.path
from copy import copy
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_FORMULA

from remarkable.config import project_root
from remarkable.converter.ecitic.util import (
    SECTION_LINE_COMMENT,
    Converter,
    CustomCell,
    DocType,
    TableItem,
    apply_formula,
    make_comment,
    nested_table_handler,
    revise_group_items,
    revise_multi_kv_items,
    revise_nested_items,
    revise_one_prefer_items,
    section_title_handler,
)


class SmallScatteredConverter(Converter):
    sections = ("基本信息", "新增基础资产情况", "购买完成后基础资产情况", "其他信息")
    template_path = os.path.join(project_root, "data/ecitic/small_scattered_template.xlsx")
    sheet_title = "“小而分散”类资产"

    @classmethod
    def load_workbook_from_json(cls, sections: list[list[list | Any]]) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = cls.sheet_title
        row_idx = 1
        for sec_title, tables in sections:
            if sec_title.get("name") not in (cls.sections[0], cls.sections[-1]):
                worksheet.cell(row_idx, 1).value = sec_title["name"]
                row_idx += 1
            for key, value in tables:
                if isinstance(key, str):
                    worksheet.cell(row_idx, 1).value = key
                    if "若有多个" in key:
                        row_idx += 1
                        worksheet.cell(row_idx, 1).value = value["text"]
                        worksheet.cell(row_idx, 1).comment = make_comment(value)
                    else:
                        worksheet.cell(row_idx, 2).value = value["text"]
                        worksheet.cell(row_idx, 2).comment = make_comment(value)
                    row_idx += 1
                elif isinstance(key, dict):
                    worksheet.cell(row_idx, 1).value = key["name"]
                    row_idx += 1
                    if key["mode"] == "vertical":
                        for t_key, t_value in value[0]:  # 特殊表: 资产池基本情况
                            worksheet.cell(row_idx, 1).value = t_key
                            worksheet.cell(row_idx, 2).value = t_value["text"]
                            worksheet.cell(row_idx, 2).comment = make_comment(t_value)
                            row_idx += 1
                    else:
                        for idx, cols in enumerate(value):
                            if idx == 0:
                                for col_idx, col in enumerate((k[0] for k in cols), 1):
                                    worksheet.cell(row_idx, col_idx).value = col
                            row_idx += 1
                            for col_idx, col in enumerate((k[1] for k in cols), 1):
                                worksheet.cell(row_idx, col_idx).value = col["text"]
                                worksheet.cell(row_idx, col_idx).comment = make_comment(col)
                        row_idx += 1
        cls.apply_style(load_workbook(cls.template_path).active, worksheet)
        return workbook

    def load_workbook(self) -> Workbook:
        sections = [
            # section 1: 基本信息
            TableItem("报告名称", ["基本情况", "报告名称"]),
            TableItem("基础资产类型", None, ""),
            TableItem("基础资产是否为“小而分散”类资产", None, "是"),
            TableItem("循环购买频率", None, "按天"),
            TableItem("循环购买报告频率", None, "季"),
            TableItem(
                "循环购买时间",
                None,
                data_handler=revise_one_prefer_items,
                sub_items=[
                    TableItem("循环购买时间", ["基本情况", "循环购买时间"]),
                    TableItem("循环购买时间", ["基本情况", "循环购买时间（报告期间）"]),
                ],
            ),
            TableItem(
                "可用于购买新增基础资产的价款总额（万元）", ["基本情况", "可用于购买新增基础资产的价款总额（万元）"]
            ),
            TableItem("可供购买的基础资产总额（万元）", ["基本情况", "可供购买的基础资产总额（万元）"]),
            TableItem("实际购买的新增基础资产总额（万元）", ["基本情况", "实际购买的新增基础资产总额（万元）"]),
            TableItem(
                "基础资产买方（若有多个，以分号分割）",
                None,
                data_handler=revise_one_prefer_items,
                sub_items=[
                    TableItem("基础资产买方（若有多个，以分号分割）", ["基本情况", "基础资产买方"]),
                    TableItem("基础资产买方（若有多个，以分号分割）", ["基本情况", "基础资产买方（段落）"]),
                ],
            ),
            TableItem(
                "基础资产卖方（若有多个，以分号分割）",
                None,
                data_handler=revise_one_prefer_items,
                sub_items=[
                    TableItem("基础资产卖方（若有多个，以分号分割）", ["基本情况", "基础资产卖方"]),
                    TableItem("基础资产卖方（若有多个，以分号分割）", ["基本情况", "基础资产卖方（段落）"]),
                ],
            ),
            TableItem("循环购买账户资金划转情况", ["基本情况", "循环购买账户资金划转情况"]),
            # section 2: 新增基础资产情况
            TableItem("新增基础资产情况", None, None, xlsx_handler=section_title_handler),
            TableItem("新增基础资产数量（笔）", ["新增基础资产情况", "新增基础资产数量（笔）"]),
            TableItem("新增基础资产金额（万元）", ["新增基础资产情况", "新增基础资产金额（万元）"]),
            TableItem(
                "新增基础资产债务人数量（个）",
                None,
                data_handler=revise_one_prefer_items,
                sub_items=[
                    TableItem("新增基础资产债务人数量（个）", ["新增基础资产情况", "新增基础资产债务人数量（个）"]),
                    TableItem("新增基础资产债务人数量（个）", ["新增基础资产情况", "新增基础资产数量（笔）"]),
                ],
            ),
            TableItem(
                "新增基础资产利率分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("基础资产利率分布（%）", ["新增基础资产利率分布", "基础资产利率分布(%)"]),
                    TableItem("未偿余额（万元）", ["新增基础资产利率分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增基础资产利率分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "新增基础资产剩余期限分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("基础资产剩余期限分布（月）", ["新增基础资产剩余期限分布", "基础资产剩余期限分布（月）"]),
                    TableItem("未偿余额（万元）", ["新增基础资产剩余期限分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增基础资产剩余期限分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "新增基础资产账期分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("基础资产账期分布（月）", ["新增基础资产账期分布", "基础资产账期分布（月）"]),
                    TableItem("未偿余额（万元）", ["新增基础资产账期分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增基础资产账期分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "新增债务人所属地区分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人所属地区", ["新增债务人所属地区分布", "债务人所属地区"]),
                    TableItem("未偿余额（万元）", ["新增债务人所属地区分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增债务人所属地区分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "新增债务人年龄分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人年龄分布（岁）", ["新增债务人年龄分布", "债务人年龄分布（岁）"]),
                    TableItem("未偿余额（万元）", ["新增债务人年龄分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["新增债务人年龄分布", "占比(%）"]),
                ],
            ),
            # section 3: 购买完成后基础资产情况
            TableItem("购买完成后基础资产情况", None, None, xlsx_handler=section_title_handler),
            TableItem(
                "基础资产变化情况",
                None,
                data_handler=revise_nested_items,
                xlsx_handler=nested_table_handler,
                sub_items=[
                    TableItem(
                        "",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", None, "基础资产数量（笔）"),
                            TableItem("基础资产金额（万元）", None, "基础资产金额（万元）"),
                            TableItem("债务人数量（个）", None, "债务人数量（个）"),
                        ],
                    ),
                    TableItem(
                        "循环购买后",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", ["基础资产变化情况-循环购买后", "基础资产数量（笔）"]),
                            TableItem("基础资产金额（万元）", ["基础资产变化情况-循环购买后", "基础资产金额（万元）"]),
                            TableItem(
                                "债务人数量（个）",
                                None,
                                data_handler=revise_one_prefer_items,
                                sub_items=[
                                    TableItem("债务人数量（个）", ["基础资产变化情况-循环购买后", "债务人数量（个）"]),
                                    TableItem(
                                        "债务人数量（个）", ["基础资产变化情况-循环购买后", "基础资产数量（笔）"]
                                    ),
                                ],
                            ),
                        ],
                    ),
                    # TODO: formula support
                    # 循环购买后 - 循环购买前 = ?
                    TableItem(
                        "报告期增减",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", None, ""),
                            TableItem("基础资产金额（万元）", None, ""),
                            TableItem("债务人数量（个）", None, ""),
                        ],
                    ),
                    TableItem(
                        "循环购买前",
                        None,
                        data_handler=revise_multi_kv_items,
                        sub_items=[
                            TableItem("基础资产数量（笔）", ["基础资产变化情况-循环购买前", "基础资产数量（笔）"]),
                            TableItem("基础资产金额（万元）", ["基础资产变化情况-循环购买前", "基础资产金额（万元）"]),
                            TableItem(
                                "债务人数量（个）",
                                None,
                                data_handler=revise_one_prefer_items,
                                sub_items=[
                                    TableItem("债务人数量（个）", ["基础资产变化情况-循环购买前", "债务人数量（个）"]),
                                    TableItem(
                                        "债务人数量（个）", ["基础资产变化情况-循环购买前", "基础资产数量（笔）"]
                                    ),
                                ],
                            ),
                        ],
                    ),
                ],
            ),
            TableItem(
                "资产池基本情况",
                None,
                data_handler=revise_group_items,
                sub_items=[
                    TableItem("项目", None, "数据情况"),
                    TableItem("单笔基础资产最高未偿本金余额", ["资产池基本情况", "单笔基础资产最高未偿本金余额"]),
                    TableItem("单笔基础资产最高年化利率", ["资产池基本情况", "单笔基础资产最高年化利率"]),
                    TableItem(
                        "单笔基础资产最高加权平均期限（月）", ["资产池基本情况", "单笔基础资产最高加权平均期限（月）"]
                    ),
                    TableItem("基础资产加权平均期限（月）", ["资产池基本情况", "基础资产加权平均期限（月）"]),
                    TableItem(
                        "单个债务人平均未偿基础资产余额（万元）",
                        ["资产池基本情况", "单个债务人平均未偿基础资产余额（万元）"],
                    ),
                    TableItem(
                        "单个债务人最高未偿基础资产余额（万元）",
                        ["资产池基本情况", "单个债务人最高未偿基础资产余额（万元）"],
                    ),
                ],
            ),
            TableItem(
                "基础资产利率分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("基础资产利率分布（%）", ["基础资产利率分布", "基础资产利率分布(%)"]),
                    TableItem("未偿余额（万元）", ["基础资产利率分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["基础资产利率分布", "占比(%）"]),
                ],
            ),
            TableItem(
                "基础资产剩余期限分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("基础资产剩余期限分布（月）", ["基础资产剩余期限分布", "基础资产剩余期限分布（月）"]),
                    TableItem("未偿余额（万元）", ["基础资产剩余期限分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["基础资产剩余期限分布", "占比（%)"]),
                ],
            ),
            TableItem(
                "基础资产账期分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("基础资产账期分布（月）", ["基础资产账期分布", "基础资产账期分布（月）"]),
                    TableItem("未偿余额（万元）", ["基础资产账期分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["基础资产账期分布", "占比（%)"]),
                ],
            ),
            TableItem(
                "债务人所属地区分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人所属地区", ["债务人所属地区分布", "债务人所属地区"]),
                    TableItem("未偿余额（万元）", ["债务人所属地区分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["债务人所属地区分布", "占比（%)"]),
                ],
            ),
            TableItem(
                "债务人年龄分布",
                None,
                data_handler=revise_multi_kv_items,
                sub_items=[
                    TableItem("债务人年龄分布（岁）", ["债务人年龄分布", "债务人年龄分布（岁）"]),
                    TableItem("未偿余额（万元）", ["债务人年龄分布", "未偿余额（万元）"]),
                    TableItem("占比（%）", ["债务人年龄分布", "占比（%)"]),
                ],
            ),
            # section 4: 其他信息
            TableItem(
                "其他与本报告事项相关且管理人认为应当披露的信息", ["其他与本报告事项相关且管理人认为应当披露的信息"]
            ),
        ]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_title
        row_idx = 0
        doc_type = DocType.NORMAL
        for section in sections:
            section.doc_type = doc_type
            section.rows = section.data_handler(self.answer_node, section)
            if section.title == "报告名称":
                # 特殊处理
                # https://gitpd.paodingai.com/cheftin/docs_scriber/-/issues/484#note_153012
                doc_type = DocType.which_type([s[-1].plain_text for s in section.rows])
            section.xlsx_handler(worksheet, section, row_idx)
            row_idx += len(section.rows)

        self.apply_style(load_workbook(self.template_path).active, worksheet)

        return workbook

    @staticmethod
    def apply_style(from_ws, to_ws):
        to_ws_rows = dict(enumerate(to_ws.rows, 1))
        rows = []
        for row_idx in range(1, to_ws.max_row + 1):
            cell = to_ws.cell(row_idx, 1)
            if CustomCell.load_comment(cell) == SECTION_LINE_COMMENT:
                rows.append([to_ws_rows[row_idx]])
            elif cell.comment is not None and rows[-1][-1][0].comment is not None:
                rows[-1].append(to_ws_rows[row_idx])
            else:
                rows.append([to_ws_rows[row_idx]])

        rows_map = dict(enumerate(rows, 1))
        for idx, items in enumerate(rows, 1):
            for row in items:
                for col_idx, cell in enumerate(row, 1):
                    from_ws_cell = from_ws.cell(idx, col_idx)
                    cell.font = copy(from_ws_cell.font)
                    cell.border = copy(from_ws_cell.border)
                    cell.fill = copy(from_ws_cell.fill)
                    cell.number_format = copy(from_ws_cell.number_format)
                    cell.protection = copy(from_ws_cell.protection)
                    cell.alignment = copy(from_ws_cell.alignment)
                    if from_ws_cell.data_type == TYPE_FORMULA and cell.comment is None:
                        apply_formula(rows_map[idx], cell, from_ws_cell.value)
                # 行高
                to_ws.row_dimensions[row[0].row].height = copy(from_ws.row_dimensions[idx].height)
        # 合并单元格
        for merged_cell in from_ws.merged_cells.ranges:
            row_idx = merged_cell.left[0][0]
            from_col_idx, to_col_idx = merged_cell.left[0][1], merged_cell.right[0][1]
            for row in rows_map[row_idx]:
                to_ws.merge_cells(
                    start_row=row[0].row, start_column=from_col_idx, end_row=row[0].row, end_column=to_col_idx
                )
        for idx in ("A", "B", "C", "D"):
            # 列宽
            to_ws.column_dimensions[idx].width = copy(from_ws.column_dimensions[idx].width)
