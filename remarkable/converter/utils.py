import csv
import logging
import re
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from functools import lru_cache

import attr
import requests
from six.moves import collections_abc

from remarkable.models.new_file import NewFile
from remarkable.pw_models.model import NewFileProject, NewMold
from remarkable.pw_models.question import NewQuestion

try:
    import zoneinfo
except ImportError:
    from backports import zoneinfo
from dateparser.search import search_dates
from openpyxl import styles
from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from remarkable.common.exceptions import PushError
from remarkable.common.storage import localstorage
from remarkable.common.util import ClassBakery, clean_txt
from remarkable.db import peewee_transaction_wrapper
from remarkable.pdfinsight.reader import PdfinsightReader
from remarkable.plugins.diff.common import gen_cache_for_diff

class_bakery = ClassBakery("web.classes.answer_converter")
p_number_only = re.compile(r"[0-9.,/-]+")
p_number_pure = re.compile(r"^[0-9.,/-]+$")
p_chinese_number = re.compile(r"[〇一二三四五六七八九零壹贰叁肆伍陆柒捌玖貮两十拾百佰千仟万萬亿億兆]+")
p_currency = re.compile(r"(?P<target>人民币|[港欧美日韩]元)")
cn_num = {
    "〇": 0,
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "零": 0,
    "壹": 1,
    "贰": 2,
    "叁": 3,
    "肆": 4,
    "伍": 5,
    "陆": 6,
    "柒": 7,
    "捌": 8,
    "玖": 9,
    "貮": 2,
    "两": 2,
}
cn_unit = {
    "天": 1,
    "日": 1,
    "周": 7,
    "月": 30,
    "十": 10,
    "拾": 10,
    "百": 100,
    "佰": 100,
    "千": 1000,
    "仟": 1000,
    "万": 10000,
    "萬": 10000,
    "亿": 100000000,
    "億": 100000000,
    "兆": 1000000000000,
}


@attr.s
class DataPack:
    file: NewFile | None = attr.ib()
    project: NewFileProject | None = attr.ib()
    mold: NewMold | None = attr.ib()
    question: NewQuestion | None = attr.ib()
    answer: dict | None = attr.ib(default=None)


class ExcelWriter:
    side = styles.Side(style="thin", color=styles.colors.BLACK)
    border = styles.Border(left=side, right=side, top=side, bottom=side)

    @classmethod
    def write(cls, data):
        raise NotImplementedError

    @classmethod
    def clean_excel_col_name(cls, name):
        return name

    @classmethod
    def write_table_row(cls, sheet: Worksheet, config: dict, data: list):
        """
        | title1 | title2 | title3|
        |  abc   |  def   |  hij  |

        row_range 包含title所在行
        """
        row_start, row_end = config["row_range"]
        if row_end == -1:
            for col in range(*config["col_range"]):
                key = sheet.cell(row_start, col).value
                key = cls.clean_excel_col_name(key)
                for index, ans in enumerate(data):
                    row = row_start + 1 + index
                    value = ans.get(key)
                    cell = sheet.cell(row, col)
                    cell.border = cls.border
                    if not value or not value.get("text"):
                        continue
                    cell.value = value["text"]

    @classmethod
    def write_table_kv(cls, sheet: Worksheet, config: dict, data: list):
        """
        | title1 | abc | title2 | def |
        | title3 | hij | title4 | klm |
        row_range, col_range 包含title所在行,列
        """
        index = config["answer_index"]
        if index + 1 > len(data):
            return

        index_data = data[index]
        for col in range(*config["col_range"]):
            for row in range(*config["row_range"]):
                cell = sheet.cell(row, col)
                key = cls.clean_excel_col_name(cell.value)
                if key:
                    value = index_data.get(key)
                    if not value or not value.get("text"):
                        continue
                    right_cell = cls.get_right_unmerged_cell(sheet, cell)
                    right_cell.value = value["text"]

    @classmethod
    def write_table_tuple(cls, sheet: Worksheet, config: dict, data: list):
        """
        |  title  | title2 | title3 |
        | title_a |  def   |  hij  |
        | title_b |  abc   |  klm  |
        row_range, col_range 包含title所在行,列
        """
        if config["answer_index"] + 1 > len(data):
            return

        index_data = data[config["answer_index"]]
        row_start, row_end = config["row_range"]
        col_start, col_end = config["col_range"]
        for row in range(row_start + 1, row_end):
            row_key = cls.clean_excel_col_name(sheet.cell(row, col_start).value)
            row_data = index_data.get(row_key)
            if not row_data:
                continue
            for col in range(col_start + 1, col_end):
                col_key = cls.clean_excel_col_name(sheet.cell(row_start, col).value)
                value = row_data.get(col_key)
                if not value or not value.get("text"):
                    continue
                cell = sheet.cell(row, col)
                cell.value = value["text"]

    @classmethod
    def get_right_unmerged_cell(cls, sheet: Worksheet, cell: Cell | MergedCell) -> Cell:
        right_cell = sheet.cell(cell.row, cell.column + 1)
        if not isinstance(right_cell, MergedCell):
            return right_cell
        return cls.get_right_unmerged_cell(sheet, right_cell)


def adjust_decimal_places(number, decimal_places=0, format_str=""):
    if not format_str:
        format_str = "0." + "0" * decimal_places
    return Decimal(float(number)).quantize(Decimal(format_str), rounding=ROUND_HALF_UP).to_eng_string()


def get_decimal_places(number):
    if "." in number:
        return len(number.split(".")[-1])
    return 0


def cn2digit(chinese_number):
    lcn = list(chinese_number)
    unit = 0  # 当前的单位
    ldig = []  # 临时数组

    while lcn:
        cndig = lcn.pop()

        if cndig in cn_unit:
            unit = cn_unit.get(cndig)
            if unit == 10000:
                ldig.append("w")  # 标示万位
                unit = 1
            elif unit == 100000000:
                ldig.append("y")  # 标示亿位
                unit = 1
            elif unit == 1000000000000:  # 标示兆位
                ldig.append("z")
                unit = 1

            continue

        dig = cn_num.get(cndig)

        if unit:
            dig = dig * unit
            unit = 0

        ldig.append(dig)

    if unit == 10:  # 处理10-19的数字
        ldig.append(10)

    ret = 0
    tmp = 0

    while ldig:
        x = ldig.pop()

        if x == "w":
            tmp *= 10000
            ret += tmp
            tmp = 0

        elif x == "y":
            tmp *= 100000000
            ret += tmp
            tmp = 0

        elif x == "z":
            tmp *= 1000000000000
            ret += tmp
            tmp = 0

        else:
            tmp += x

    ret += tmp
    return ret


def split_number_and_unit(text):
    res = re.search(r"(?P<val>-?\d+\.?\d*)(?P<unit>\D{0,4})$", text.replace(",", ""))
    if res:
        return res.group("val"), res.group("unit")
    return None, None


def convert_unit(val, from_unit="", to_unit="", decimal_places=0):
    def get_multi(unit):
        multipliers = {"十": 10, "拾": 10, "百": 100, "佰": 100, "千": 1000, "仟": 1000, "万": 10000, "亿": 100000000}
        denominator_multipliers = {"吨": 1000, "千克": 1, "公斤": 1, "斤": 0.5}
        multi = Decimal(1)
        if "/" in unit:
            numerator_unit, denominator_unit = unit.split("/")
        else:
            numerator_unit = unit
            denominator_unit = ""

        for word, _multi in multipliers.items():
            if word in numerator_unit:
                multi *= Decimal(_multi)
        for word, _multi in denominator_multipliers.items():
            if word in denominator_unit:
                multi /= Decimal(_multi)

        return multi

    val_str, unit_in_val = split_number_and_unit(clean_txt(val))
    if not val_str:
        return None

    # 就近取单位
    from_multi = get_multi(unit_in_val or from_unit)
    to_multi = get_multi(to_unit)

    # 保留计算精度, 等到按规定格式输出时再做四舍五入
    val = Decimal(val_str) * from_multi / to_multi
    return adjust_decimal_places(val, decimal_places, str(val))


def get_currency_unit(unit):
    if not unit:
        return None
    match = p_currency.search(unit)
    if match:
        return match.groupdict()["target"]
    return None


def keep_number_only(text, key=""):
    if not key and p_number_pure.search(text):
        return text
    match = p_number_only.search(text)
    if match:
        number = match.group()
        if "%" in text or key:
            decimal_places = get_decimal_places(number)
            number = adjust_decimal_places(Decimal(number) / 100, decimal_places=decimal_places + 2)
    else:
        number = None
    return number


def chinese_number_convert(text):
    if p_number_pure.search(text):
        return text
    match = p_chinese_number.search(text)
    if match:
        text = match.group()
        text = cn2digit(text)
    else:
        text = None
    return text


def cn2en_date(text):
    text = text.replace("号", "日")
    if "十" in text:
        # "十"前后有数字, 去掉"十"
        text = re.sub(
            r"([一二三四五六七八九壹贰叁肆伍陆柒捌玖貮])十([一二三四五六七八九壹贰叁肆伍陆柒捌玖貮].*)", r"\1\2", text
        )
        # "十"前后无数字, 添个"零"
        text = re.sub(r"(^十)([^一二三四五六七八九壹贰叁肆伍陆柒捌玖貮].*)", r"\1零\2", text)
        text = re.sub(
            r"([^一二三四五六七八九壹贰叁肆伍陆柒捌玖貮])(十)([^一二三四五六七八九壹贰叁肆伍陆柒捌玖貮].*)",
            r"\1\2零\3",
            text,
        )
        text = text.replace("十", "1")
    return "".join(str(cn_num.get(t, t)) for t in text)


@lru_cache()
def date_from_text(date_str: str, languages=("en", "zh", "zh-Hans")) -> datetime | None:
    time_zone = zoneinfo.ZoneInfo("Asia/Shanghai")
    relative_base = datetime.now(tz=time_zone)
    settings = {
        "RELATIVE_BASE": relative_base,
        "TIMEZONE": str(time_zone),
    }
    date_list = search_dates(cn2en_date(date_str), languages=languages, settings=settings)
    if not date_list:
        return None
    return date_list[0][-1]


def push(dst_url, data):
    rsp = requests.post(url=dst_url, json=data, timeout=30)
    if not rsp.ok:
        raise PushError(rsp.json())


def csv_reader(csv_path, skip_rows=1):
    with open(csv_path, "r", newline="") as file_iter:
        reader = csv.reader(file_iter)
        for idx, row in enumerate(reader):
            if idx >= skip_rows:
                yield row


async def prepare_data(qid) -> DataPack:
    question = await NewQuestion.find_by_id(qid)
    file = await NewFile.find_by_id(question.fid)
    project = await NewFileProject.find_by_id(file.pid)
    mold = await NewMold.find_by_id(question.mold)
    return DataPack(file, project, mold, question)


def get_answer_workshop(mold_name):
    workshop = class_bakery.get_class(mold_name, class_bakery)
    return workshop


async def call_workshop(meta_data, debug):
    workshop = get_answer_workshop(meta_data.mold.name)

    if not meta_data.answer:
        logging.warning(f"answer is None, qid: {meta_data.question.id}")
        return

    if workshop:
        await workshop(meta_data, debug).work()
    else:
        logging.warning(f"converter class not found: {class_bakery.config_entry}")


@peewee_transaction_wrapper
async def push_answer_to_remote(qid, answer_type="ai", debug=False):
    """
    answer_type = 'ai' or 'manual'
    """
    logging.info(f"start to push preset_answer, qid:{qid}, {answer_type}")
    meta_data = await prepare_data(qid)
    meta_data.answer = meta_data.question.preset_answer if answer_type == "ai" else meta_data.question.answer
    await call_workshop(meta_data, debug)


@peewee_transaction_wrapper
async def generate_customer_answer(qid, debug=False):
    meta_data = await prepare_data(qid)
    logging.info(f"start to generate_customer_answer, {qid=}, fid: {meta_data.question.fid}")
    meta_data.answer = meta_data.question.answer
    await call_workshop(meta_data, debug)


async def generate_cache_for_diff(qid):
    question = await NewQuestion.find_by_id(qid)
    answer = question.answer
    file = await NewFile.find_by_qid(qid)
    reader = PdfinsightReader(localstorage.mount(file.pdfinsight_path()))
    gen_cache_for_diff(answer, reader)


def flatten_dict(item, prefix="", sep="-", keep_index=False):
    def _take_prefix(key, value, pre):
        if pre:
            yield from flatten_dict(value, f"{pre}{sep}{key}", sep, keep_index=keep_index)
        else:
            yield from flatten_dict(value, str(key), sep, keep_index=keep_index)

    if isinstance(item, dict):
        for key, value in item.items():
            if isinstance(value, str) or not isinstance(value, collections_abc.Iterable):
                f_prefix = f"{prefix}{sep}" if prefix else ""
                f_index = ":0" if keep_index else ""
                yield f"{f_prefix}{key}{f_index}", value
            elif isinstance(value, dict):
                yield from _take_prefix(key, value, prefix)
            elif isinstance(value, list):
                if value:
                    for i, data in enumerate(value):
                        f_index = f":{i}" if keep_index else ""
                        yield from _take_prefix(f"{key}{f_index}", data, prefix)
                else:
                    f_prefix = f"{prefix}{sep}" if prefix else ""
                    yield f"{f_prefix}{key}", ""
            else:
                pass
    else:
        pass
