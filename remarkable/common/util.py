from __future__ import annotations

import asyncio
import contextlib
import functools
import gc
import hashlib
import json
import logging
import os
import random
import re
import subprocess
import sys
import time
import zipfile
from collections import OrderedDict
from copy import deepcopy
from datetime import datetime
from importlib import import_module
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterator, Type, TypeVar

import chardet
import filetype
import numpy as np
import redis_lock
from aiofile import async_open
from caio import python_aio_asyncio
from msgspec.json import decode
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfparser import PDFParser
from pdfminer.pdftypes import resolve1
from pdfparser.pdftools.pdf_element import extract_paras_by_outline
from pdfparser.pdftools.pdf_util import PDFUtil
from utensils.archive_reader import fallback_decode

from remarkable.common.constants import AnswerType, HistoryAction, Language
from remarkable.common.exceptions import CustomError, ShellCmdError
from remarkable.config import IS_TEST_ENV, _config, get_config
from remarkable.db import init_rdb
from remarkable.infrastructure.mattermost import MMPoster

logger = logging.getLogger(__name__)
DATE_PATTERN = re.compile(r"(\d+)年度?(\d+月份?)?(\d+日)?")
P_WHITE = re.compile(r"\s")
need_secure_cookie = (get_config("web.scheme") or "http").lower() == "https"
T = TypeVar("T")

# 科创板招股说明书信息抽取POC导出模板
template_2 = OrderedDict(
    {
        "B-1-1": {
            "label": "发行人基本情况",
            "src_label": "发行人情况",
            "hyphen": "-",
            "record": [
                {
                    "发行人名称": {"type": "text", "data": []},
                    "成立日期": {"type": "text", "data": []},
                    "行业分类": {"type": "text", "data": []},
                    "在其他交易场所(申请)挂牌或上市的情况": {"type": "text", "data": []},
                }
            ],
        },
        "B-1-2": {
            "label": "本次发行的有关中介机构",
            "src_label": "中介机构",
            "hyphen": "-",
            "record": [
                {
                    "保荐人": {"type": "text", "data": []},
                    "主承销商": {"type": "text", "data": []},
                    "发行人律师": {"type": "text", "data": []},
                    "其他承销机构": {"type": "text", "data": []},
                }
            ],
        },
        "B-2-1": {
            "label": "本次发行的基本情况",
            "src_label": "发行情况",
            "hyphen": "-",
            "record": [
                {
                    "每股面值": {"type": "text", "data": []},
                    "发行股数": {"type": "text", "data": []},
                    "发行股数占发行后总股本比例": {"type": "text", "data": []},
                    "发行后总股本": {"type": "text", "data": []},
                    "募集资金总额": {"type": "text", "data": []},
                    "募集资金净额": {"type": "text", "data": []},
                    "募集资金投资项目": {"type": "text", "data": []},
                    "发行费用概算": {"type": "text", "data": []},
                }
            ],
        },
        "B-3": {
            "label": "发行人报告期的主要财务数据和财务指标",
            "src_label": "主要财务数据和财务指标",
            "record": [
                {
                    "时间": {"type": "text", "data": []},
                    "资产总额": {"type": "text", "data": []},
                    "归属于母公司所有者权益": {"type": "text", "data": []},
                }
            ],
        },
        "B-6-1": {
            "label": "发行人选择的具体上市标准",
            "src_label": "上市标准",
            "hyphen": "",
            "record": [{"条款编号": {"type": "text", "data": []}}],
        },
        "B-6-2": {
            "label": "发行人选择的具体上市标准",
            "src_label": "上市标准",
            "hyphen": "",
            "record": [{"具体标准内容": {"src_label": "具体内容", "type": "text", "data": []}}],
        },
        "C-2-1": {
            "label": "中介机构_保荐人(主承销商)",
            "src_label": "保荐人",
            "hyphen": "-",
            "record": [
                {
                    "机构名称": {"type": "text", "data": []},
                    "发行代表人": {"src_label": "法定代表人", "type": "text", "data": []},
                }
            ],
        },
        "E-8-1": {
            "label": "发行人股东情况_实际控制人",
            "src_label": "实际控制人",
            "hyphen": "-",
            "record": [{"实际控制人": {"src_label": "姓名", "type": "text", "data": []}}],
        },
        "E-8-5": {
            "label": "发行人股东情况_控股股东/实际控制人股份质押情况",
            "src_label": "",
            "hyphen": "",
            "record": [{"控股股东/实际控制人股份是否存在质押情况": {"type": "text", "data": []}}],
        },
        "E-9-2": {
            "label": "发行人股东情况_前十名股东",
            "src_label": "前十名股东",
            "record": [
                {
                    "股东姓名/名称": {"type": "text", "data": []},
                    "持股数量": {"type": "text", "data": []},
                    "持股比例": {"type": "text", "data": []},
                }
            ],
        },
        "E-9-6": {"label": "发行人股东情况_股东关系", "record": [{"股东关系": {"type": "text", "data": []}}]},
        "E-10-1": {
            "label": "发行人股东情况_董事会成员",
            "src_label": "董事会成员",
            "record": [
                {
                    "姓名": {"type": "text", "data": []},
                    "任期": {"type": "text", "data": []},
                    "简历": {"type": "text", "data": []},
                }
            ],
        },
        "F-4-2": {
            "label": "业务与技术_前五供应商",
            "src_label": "前五供应商",
            "record": [
                {
                    "时间": {"type": "text", "data": []},
                    "供应商名称": {"type": "text", "data": []},
                    "采购额": {"type": "text", "data": []},
                    "货币单位": {"type": "text", "data": []},
                }
            ],
        },
        "G-3": {
            "label": "公司治理与独立性_发行人协议控制架构情况",
            "record": [{"发行人协议控制架构情况": {"type": "text", "data": []}}],
        },
        "H-1-3": {
            "label": "财务会计信息_合并资产负债表",
            "src_label": "合并资产负债表",
            "record": [
                {
                    "时间": {"type": "text", "data": []},
                    "货币基金(流动资产)": {"src_label": "货币资金", "type": "text", "data": []},
                    "长期股权投资(非流动资产)": {"src_label": "长期股权投资", "type": "text", "data": []},
                    "非流动资产合计(非流动资产)": {"src_label": "非流动资产合计", "type": "text", "data": []},
                    "资产总计": {"type": "text", "data": []},
                }
            ],
        },
        "H-2": {"label": "财务会计信息_审计意见", "record": [{"审计意见": {"type": "text", "data": []}}]},
        "H-10-2": {
            "label": "财务会计信息_营业收入分区域分析",
            "src_label": "营业收入分区域分析",
            "record": [
                {
                    "时间": {"type": "text", "data": []},
                    "地区": {"type": "text", "data": []},
                    "收入": {"type": "text", "data": []},
                    "单位": {"src_label": "货币单位", "type": "text", "data": []},
                    "占比": {"type": "text", "data": []},
                }
            ],
        },
        "I-1-1": {
            "label": "募集资金_募集资金总量及使用情况",
            "src_label": "募集资金总量及使用情况",
            "record": [
                {
                    "货币单位": {"type": "text", "data": []},
                    "项目名称": {"type": "text", "data": []},
                    "总投资额": {"src_label": "投资总额", "type": "text", "data": []},
                    "募集资金投资额": {"type": "text", "data": []},
                    "审批文号": {"type": "text", "data": []},
                }
            ],
        },
        "K-3": {
            "label": "其他重要事项_重大诉讼",
            "record": [
                {
                    "是否有重大诉讼": {"src_label": "重大诉讼情况", "type": "enum", "data": []},
                    "重大诉讼情况": {"type": "text", "data": []},
                }
            ],
        },
    }
)

template_3 = OrderedDict(
    {
        "A": {"label": "证券代码", "record": [{"证券代码": {"type": "text", "data": []}}]},
        "B": {"label": "证券简称", "record": [{"证券简称": {"type": "text", "data": []}}]},
        "C": {"label": "公告编号", "record": [{"公告编号": {"type": "text", "data": []}}]},
        "D": {"label": "变更后的证券简称", "record": [{"变更后的证券简称": {"type": "text", "data": []}}]},
        "E": {"label": "变更日期", "record": [{"变更日期": {"type": "text", "data": []}}]},
        "F": {"label": "公告日", "record": [{"公告日": {"type": "text", "data": []}}]},
    }
)

templates = {2: template_2, 3: template_3}


class Singleton:
    _instance = None
    _inited = False

    def __new__(cls, *args, **kwargs):
        if not cls._instance:
            cls._instance = super(Singleton, cls).__new__(cls)
        return cls._instance

    def _pop(
        self,
    ):
        pass


def answer_type_to_history_action_type(answer_type):
    return {
        AnswerType.USER_DO.value: HistoryAction.SUBMIT_ANSWER.value,
        AnswerType.ADMIN_DO_1.value: HistoryAction.SUBMIT_ANSWER.value,
        AnswerType.ADMIN_DO_2.value: HistoryAction.SUBMIT_ANSWER.value,
        AnswerType.ADMIN_VERIFY.value: HistoryAction.ADMIN_VERIFY.value,
        AnswerType.ADMIN_JUDGE.value: HistoryAction.ADMIN_JUDGE.value,
    }[answer_type]


def md5(_str):
    _md5 = hashlib.md5()
    _md5.update(_str)
    return _md5.hexdigest()


@functools.lru_cache(1000)
def clean_txt(text, language=None, remove_cn_text=False, lstrip=False):
    if not language:
        language = get_config("client.content_language") or "zh_CN"
    if language == Language.ZH_CN.value:
        text = re.sub(r"\s+", "", text)
    elif language == Language.EN_US.value:
        # 英文内容换行需要用空格链接,且不能去空格
        text = re.sub(r"[\r\t\f\v\n]+", " ", text)
        text = re.sub(r"\s{2,}", " ", text)
        if lstrip:
            text = text.lstrip()
        else:
            text = text.strip()

    if remove_cn_text:
        text = re.sub(r"[\u4e00-\u9fa5]", "", text)
    return text


def is_space(string):
    return P_WHITE.sub("", string) == ""


def get_num_from_chars(chars, char_pattern, clear=""):
    if not char_pattern:
        return chars
    if isinstance(char_pattern, str):
        char_pattern = re.compile(char_pattern)
    if clear:
        chars = re.sub(clear, "", chars)
    matched = char_pattern.search(chars)
    if matched:
        idx = matched.re.groupindex["target"]
        chars = chars[matched.start(idx) : matched.end(idx)]
        return chars
    return ""


def index_in_space_string(text, rng, *, is_cn=None):
    if is_cn is None:
        is_cn = (get_config("client.content_language") or Language.ZH_CN.value) == Language.ZH_CN.value
    _start, _end = rng
    length = _end - _start
    _pre, _text = text[:_start], text[_start:]

    language = Language.ZH_CN.value if is_cn else Language.EN_US.value
    delta_start = _start - len(clean_txt(_pre, lstrip=True, language=language))

    start = start_char = None
    if _pre:
        is_last_char_space = is_space(_pre[-1])
    else:
        is_last_char_space = P_WHITE.match(_text)
        if not is_cn and is_last_char_space:
            _start = len(_text) - len(_text.lstrip())
            _text = _text.lstrip()
    for idx, sp_char in enumerate(_text):
        if (is_last_char_space or is_cn) and is_space(
            sp_char
        ):  # 英文环境,如果上一个字符是空格，且当前字符是空格，则跳过
            continue
        is_last_char_space = is_space(sp_char)
        if delta_start > 0:
            delta_start -= 1
            continue
        if start_char is None:
            start_char = sp_char
            start = _start + idx
            if length == 0:
                return start, start

        length -= 1
        if length == 0:
            return start, (_start + idx + 1)
    return -1, -1


def cut_text(text, side="right", max_groups=1, max_chars=20):
    texts_cut = clean_txt(text)
    res = ""
    if side == "ends":
        if len(texts_cut) <= 20:
            res = texts_cut
        else:
            left_part = texts_cut[:10]
            right_part = texts_cut[len(left_part) - max_chars :]
            res = left_part + right_part
    elif side == "left":
        # 不应该把标点符号去掉，可能本身也是特征
        # 取目标左侧的一部分文字，先去掉末尾句号/分号
        # groups = re.split(r'[。.;；，,]', re.sub(r'[。.;；]$', '', texts_cut))
        # res = "".join(groups[-max_groups:])
        res = texts_cut[-max_chars:]
    else:
        # 取目标右侧的一部分文字，先去掉句首逗号/分号
        # groups = re.split(r'[。.;；，,]', re.sub(r'^[;；，,]', '', texts_cut))
        # res = "".join(groups[:max_groups])
        res = texts_cut[:max_chars]

    return res


def cut_words(text, side="right", max_words=20):
    import rjieba as jieba

    words = list(jieba.cut(text))
    res = []
    if side == "ends":
        if len(words) <= max_words:
            res = words
        else:
            left_part = words[: int(max_words / 2)]
            right_part = words[len(left_part) - max_words :]
            res = left_part + right_part
    elif side == "left":
        res = words[-max_words:]
    else:
        # 取目标右侧的一部分文字，先去掉句首逗号/分号
        res = words[:max_words]

    return res


def is_none_or_whitespace(text):
    if not text:
        return True
    return re.match(r"^\s+$", text)


def group_cells(cells):
    cells = deepcopy(cells)
    _row = {}
    _col = {}
    for idx, cell in cells.items():
        row, col = idx.split("_")
        _col.setdefault(col, {})[row] = cell
        _row.setdefault(row, {})[col] = cell

    cells_by_row = OrderedDict()
    for k in sorted(_row, key=int):
        cells_by_row[k] = OrderedDict()
        for k_1 in sorted(_row[k], key=int):
            cells_by_row[k][k_1] = _row[k][k_1]

    cells_by_col = OrderedDict()
    for k in sorted(_col, key=int):
        cells_by_col[k] = OrderedDict()
        for k_1 in sorted(_col[k], key=int):
            cells_by_col[k][k_1] = _col[k][k_1]
    return cells_by_row, cells_by_col


def filter_tree(tree_id, data, tree_s):
    tree_s.add(tree_id)
    if tree_id in data:
        for sub_tree in data[tree_id]["children"]:
            filter_tree(sub_tree["id"], data, tree_s)


def str2key(key_str):
    key_path = [i.split(":") for i in json.loads(key_str)[1:]]
    label = "".join([key for key, idx in key_path])
    idx = key_path[-2][1] if len(key_path) > 1 else 0
    return label, idx


def pick_answer(answers, label, dst_type):
    ret = {}
    for answer in answers:
        _label, _idx = str2key(answer["key"])
        if dst_type == "enum" and _label == label:
            ret.setdefault(_idx, []).append(answer["value"])
            break
        if _label == label:
            for item in answer["data"]:
                texts = []
                for box in item["boxes"]:
                    texts.append(box["text"].strip())
                ret.setdefault(_idx, []).append("".join(texts))
    return ret


def export_answer(answer, schema_id):
    temps = deepcopy(templates[int(schema_id)])
    for _, temp in temps.items():
        record_template = temp["record"][0]
        records = {}
        for key in record_template:
            label = temp.get("src_label", "") + temp.get("hyphen", "") + record_template[key].get("src_label", key)
            ans = pick_answer(answer.get("userAnswer", {}).get("items", []), label, record_template[key].get("type"))
            if key.endswith("单位") and len(ans) == 1:
                for i in range(1, len(records)):
                    ans.setdefault(str(i), ans[0])
            for idx, _ans in ans.items():
                item = records.setdefault(idx, deepcopy(record_template))
                item[key]["data"] = _ans
        temp["record"] = list(records.values())
    # 删除不必要的key
    for _, temp in temps.items():
        try:
            temp.pop("src_label")
            temp.pop("hyphen")
        except KeyError:
            pass
        for item in temp["record"]:
            for key in item:
                item[key] = item[key]["data"]  # 用data填充对应key的值, 扔掉type内容
    return temps


def calc_cost_time(start, end):
    """计算处理时间, 如果间隔时间过短或过长, 随机一个4~30min的时间"""
    res = end - start
    res = res if 1800 > res > 240 else random.randint(240, 1800)
    return "{}s".format(res)


def rate_limit(addr, prefix="rate_limit"):
    """限制接口访问频率
    TODO: 目前只应用于登录接口限制, 且IP限制存在NAT误伤问题
    :param addr: ip addr
    :param prefix: db key prefix
    :return: int: 0, 表示未达到限制次数; >0, 表示达到限制次数, 返回值表示封禁剩余时间(s)
    """
    lua = """
        local times = redis.call('incr',KEYS[1])
        if times == 1 then
            redis.call('expire',KEYS[1], ARGV[1])
        end
        if times > tonumber(ARGV[2]) then
            return redis.call('ttl',KEYS[1])
        end
        return 0
    """
    key = f"{get_config('app.app_id')}:{prefix}:{addr}"
    return init_rdb().register_script(lua)(
        keys=[key], args=[get_config("app.rate_limit_expire", 60), get_config("app.max_fail_count", 10)]
    )


def is_equal_enum_val(item1, item2):
    """
    answer_item的枚举值是否相等，兼容 null/string/list 三种情况
    """

    def to_set(val):
        res = set()
        if not isinstance(val, list):
            res.add(val)
        else:
            for x in val:
                res.add(x)
        return res

    if not ("value" in item1 and "value" in item2):
        return False

    val1 = to_set(item1["value"])
    val2 = to_set(item2["value"])

    return val1 == val2


def is_number_str(data):
    try:
        float(data)
    except ValueError:
        return False
    return True


def is_valid_answer(answer):
    # if not answer or not isinstance(answer, dict):
    #     return False
    # return bool(answer.get('userAnswer', {}).get('items'))
    return answer and isinstance(answer, dict)


def is_aim_element(elt, aim_types=None, neg_patterns=None):
    aim_types = aim_types or []
    if aim_types and elt.get("class") not in aim_types:
        return False
    text = elt.get("text", "")  # todo: text of table
    neg_patterns = neg_patterns or []
    for neg_pattern in neg_patterns:
        if re.search(neg_pattern, text):
            return False
    return True


def generate_timestamp():
    return int(time.time())


def loop_wrapper(func):
    @functools.wraps(func)
    def run_in_loop(*args, **kwargs):
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            # Create a new event loop instance and set it as current loop in new sub-thread process
            loop = asyncio.new_event_loop()
        else:
            if loop.is_running():
                # Create a new event loop object if current loop is running(possible multiprocessing env)
                loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(func(*args, **kwargs))

    return run_in_loop


def md5sum(abs_path):
    hash_md5 = hashlib.md5()
    with open(abs_path, "rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def name2driver(name):
    if name.lower() == "mysql":
        return "mysql+pymysql"
    if name.lower() == "postgresql":
        return "postgresql+psycopg2"
    if name.lower() == "oracle":
        return "oracle"
    return name


def md5json(data, default=None):
    if not isinstance(data, (str, dict)):
        raise TypeError
    if isinstance(data, str):
        data = json.loads(data)
    return hashlib.md5(json.dumps(data, sort_keys=True, default=default).encode()).hexdigest()


def subprocess_exec(command, timeout=None):
    """
    Execute a shell command with real-time output to logging while capturing the result.

    Args:
        command (str): The shell command to execute
        timeout (float, optional): Timeout in seconds for command execution

    Returns:
        str: The stdout output from the command

    Raises:
        ShellCmdError: If the command produces stderr output
        subprocess.TimeoutExpired: If the command times out
    """
    logger.info(f'running command: "{command}"')

    import threading
    from queue import Empty, Queue

    # Create a separate logger for command output to avoid mixing with application logs
    cmd_logger = logging.getLogger(f"{__name__}.subprocess_output")

    def read_stream(stream, queue, stream_name):
        """Read from stream and put lines into queue."""
        try:
            for line in iter(stream.readline, ""):
                if line:
                    queue.put((stream_name, line.rstrip("\n\r")))
            queue.put((stream_name, None))  # Signal end of stream
        except Exception as e:
            queue.put((stream_name, f"Error reading {stream_name}: {e}"))

    with subprocess.Popen(
        command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, universal_newlines=True, bufsize=1
    ) as cmd_process:
        # Create queues and threads for reading stdout and stderr
        output_queue = Queue()
        stdout_thread = threading.Thread(target=read_stream, args=(cmd_process.stdout, output_queue, "stdout"))
        stderr_thread = threading.Thread(target=read_stream, args=(cmd_process.stderr, output_queue, "stderr"))

        stdout_thread.daemon = True
        stderr_thread.daemon = True
        stdout_thread.start()
        stderr_thread.start()

        stdout_lines = []
        stderr_lines = []
        streams_ended = set()

        try:
            # Read output in real-time
            while len(streams_ended) < 2:  # Wait for both stdout and stderr to end
                try:
                    # Check if process has finished and timeout
                    if cmd_process.poll() is not None:
                        # Process finished, read remaining output with short timeout
                        stream_name, line = output_queue.get(timeout=0.1)
                    else:
                        # Process still running, check timeout
                        queue_timeout = 0.1 if timeout is None else min(0.1, timeout)
                        stream_name, line = output_queue.get(timeout=queue_timeout)

                    if line is None:
                        # Stream ended
                        streams_ended.add(stream_name)
                        continue

                    if stream_name == "stdout":
                        stdout_lines.append(line)
                        # Log stdout as INFO level
                        cmd_logger.info(f"[STDOUT] {line}")
                    elif stream_name == "stderr":
                        stderr_lines.append(line)
                        # Log stderr as WARNING level
                        cmd_logger.warning(f"[STDERR] {line}")

                except Empty:
                    # Check if we should timeout
                    if timeout is not None and cmd_process.poll() is None:
                        try:
                            cmd_process.wait(timeout=0.1)
                        except subprocess.TimeoutExpired:
                            cmd_process.kill()
                            cmd_process.wait()
                            raise subprocess.TimeoutExpired(command, timeout) from None
                    continue

            # Wait for process to complete
            cmd_process.wait(timeout=timeout)

        except subprocess.TimeoutExpired:
            cmd_process.kill()
            cmd_process.wait()
            raise
        finally:
            # Ensure threads finish
            stdout_thread.join(timeout=1)
            stderr_thread.join(timeout=1)

    # Join all stderr output
    err_output = "\n".join(stderr_lines) if stderr_lines else ""
    if err_output:
        raise ShellCmdError(err_output)

    # Join all stdout output and return
    return "\n".join(stdout_lines) if stdout_lines else ""


def read_zip_first_file(zip_path, *, msgspec_type: Type[T] | None = None) -> str | T:  # noqa: UP006
    # start = time.time()
    thin_path = zip_path + ".thin"
    if os.path.exists(thin_path):
        zip_path = thin_path
    with zipfile.ZipFile(zip_path, "r") as fzip:
        data = fzip.read(fzip.namelist()[0])
        # logging.info("read %s file cost %.2fs", zip_path, time.time() - start)
        if msgspec_type is None:
            return data.decode("utf-8") if isinstance(data, bytes) else data
        return decode(data, type=msgspec_type)


def outline_to_box(outline: list | tuple) -> dict[str, float]:
    box = {
        "box_top": outline[1],
        "box_right": outline[2],
        "box_bottom": outline[3],
        "box_left": outline[0],
    }
    return box


def box_to_outline(box: dict[str, float]) -> tuple[float, float, float, float]:
    outline = (box["box_left"], box["box_top"], box["box_right"], box["box_bottom"])
    return outline


class ClassBakery:
    CLASSNAME_OVER_CONFIG: str | None = None
    cls_cache = {}
    config_entry = "web.classes.answer_predictor"

    def __init__(self, config_entry=None):
        self.cls_cache = {}
        if config_entry:
            self.config_entry = config_entry

    @classmethod
    def get_class_fullname(cls, schema_name: str, bakery: "ClassBakery" | None = None) -> str | None:
        clz = bakery if isinstance(bakery, ClassBakery) else cls
        cls_path = cls.CLASSNAME_OVER_CONFIG or get_config(clz.config_entry)
        if isinstance(cls_path, dict):
            ret = cls_path.get(schema_name)
            if ret:
                return ret
            return cls_path.get("default")
        return cls_path

    @classmethod
    def get_class(cls, schema_name: str, bakery: "ClassBakery" | None = None) -> type | None:
        clz = bakery if isinstance(bakery, ClassBakery) else cls
        if schema_name not in clz.cls_cache:
            class_fullname = clz.get_class_fullname(schema_name, clz)
            if not class_fullname:
                return None
            clazz = import_class_by_path(class_fullname)
            if clazz:
                clz.cls_cache[schema_name] = clazz
        return clz.cls_cache.get(schema_name)


def import_class_by_path(module_path: str) -> type | None:
    last_dot_idx = module_path.rindex(".")
    module_path, cls_name = module_path[:last_dot_idx], module_path[last_dot_idx + 1 :]
    try:
        module = import_module(module_path)
        return getattr(module, cls_name, None)
    except (ImportError, ModuleNotFoundError) as e:
        logger.debug(e)
        return None


def params_dumper(obj):
    from remarkable.pw_models.base import BaseModel

    if isinstance(obj, BaseModel):
        return obj.to_dict(only=(BaseModel.id,)) or None
    raise TypeError(f"{obj} is not serializable.")


def release_lock_keys():
    rdb = init_rdb()
    for key in rdb.scan_iter(match="lock:*"):
        rdb.delete(key)


def run_singleton_task(func: Callable, *args, **kwargs) -> tuple[bool, redis_lock.Lock | None]:  # noqa: UP006
    lock_expired = kwargs.pop("lock_expired", random.randint(590, 610))
    lock_key = kwargs.pop(
        "lock_key", f"{func.__name__}:{md5json({'args': args, 'kwargs': kwargs}, default=params_dumper)}"
    )
    if not IS_TEST_ENV:
        lock = redis_lock.Lock(init_rdb(), lock_key, expire=lock_expired)
    else:
        lock = None
    if lock is None or lock.acquire(blocking=False):
        _ = func.delay(*args, **kwargs) if hasattr(func, "delay") else func(*args, **kwargs)
        return True, lock
    logger.warning(f'Someone else has the lock: "{func.__name__}", skip this task.')
    return False, lock


async def arun_singleton_task(func: Callable, *args, **kwargs) -> tuple[bool, redis_lock.Lock | None]:  # noqa: UP006
    lock_expired = kwargs.pop("lock_expired", random.randint(590, 610))
    lock_key = kwargs.pop(
        "lock_key", f"{func.__name__}:{md5json({'args': args, 'kwargs': kwargs}, default=params_dumper)}"
    )
    if not IS_TEST_ENV:
        lock = redis_lock.Lock(init_rdb(), lock_key, expire=lock_expired)
    else:
        lock = None
    if lock is None or lock.acquire(blocking=False):
        await func(*args, **kwargs)
        return True, lock
    logger.warning(f'Someone else has the lock: "{func.__name__}", skip this task.')
    return False, lock


def release_parse_file_lock(checksum):
    init_rdb().delete(f"lock:convert_or_parse_file:{checksum}")


def clear_caches():
    # All objects cleared
    for obj in (i for i in gc.get_objects() if isinstance(i, functools._lru_cache_wrapper)):
        obj.cache_clear()
    _config.reload()  # We need reload our own config(especially the new ENV variable) before os.fork
    gc.collect()


def box_in_box(element_outline, box_outline):
    """outline:
    (left, top, right, bottom)
    or
    {"box_bottom": bottom, "box_left": left, ...}
    """

    def normalize_outline(outline):
        if isinstance(outline, (tuple, list)):
            return outline
        if isinstance(outline, dict):
            return (outline["box_left"], outline["box_top"], outline["box_right"], outline["box_bottom"])
        raise ValueError("not support input outline")

    element_outline = normalize_outline(element_outline)
    box_outline = normalize_outline(box_outline)

    x_in_box = box_outline[0] <= (element_outline[2] + element_outline[0]) / 2 <= box_outline[2]
    y_in_box = box_outline[1] <= (element_outline[3] + element_outline[1]) / 2 <= box_outline[3]

    return x_in_box and y_in_box


def chars_in_box_by_center(box, page=None, texts=None, with_white_chars=False):
    chars = []

    def extend_chars(candidate_chars):
        for char in candidate_chars:
            if PDFUtil.is_box_in_box_by_center(char["box"], box):
                chars.append(char)

    if page:
        for text in page.get("texts", []):
            if text["type"] == "LINE":
                continue
            if text["box"][3] < box[1]:
                continue
            if text["box"][1] > box[3]:
                if text["box"][1] > box[3] + max(12, text["box"][3] - text["box"][1]) * 2:
                    break
                continue
            extend_chars(text["chars"])
            if with_white_chars:
                extend_chars(text.get("white_chars", []))
        chars = sorted(chars, key=functools.cmp_to_key(PDFUtil.compare_box_item))
    elif texts:
        extend_chars(texts)
        if not with_white_chars:
            chars = [char for char in chars if not P_WHITE.match(char)]
    return chars


def count_pdf_pages(data: bytes) -> int:
    """
    :param data: pdf data
    :return: page count
    """
    parser = PDFParser(BytesIO(data))
    doc = PDFDocument(parser)
    parser.set_document(doc)
    pages = resolve1(doc.catalog["Pages"])
    return pages.get("Count", 0)


def kmeans(nums, clusters=2):
    """k-means聚类算法
    clusters      - 指定分簇数量
    nums          - ndarray(m, n)，m个样本的数据集，每个样本n个属性值
    将初步定位的结果按照分数分簇
    nums 应该是有序的
    """
    sample_size, attrs = nums.shape  # sample_size：样本数量，attrs：每个样本的属性值个数
    result = np.ones(sample_size, dtype=np.int64)  # sample_size个样本的聚类结果
    cores = [nums[0], nums[-1]]  # 选择第一个和最后一个元素块当做质心

    while True:  # 迭代计算
        square = np.square(np.repeat(nums, clusters, axis=0).reshape(sample_size, clusters, attrs) - cores)
        distance = np.sqrt(
            np.sum(square, axis=2)
        )  # ndarray(sample_size, clusters)，每个样本距离k个质心的距离，共有sample_size行
        index_min = np.argmin(distance, axis=1)  # 每个样本距离最近的质心索引序号

        if (index_min == result).all():  # 如果样本聚类没有改变
            return result  # 则返回聚类结果

        result[:] = index_min  # 重新分类
        for cluster in range(clusters):  # 遍历质心集
            items = nums[result == cluster]  # 找出对应当前质心的子样本集
            cores[cluster] = np.mean(items, axis=0)  # 以子样本集的均值作为当前质心的位置


def extract_text_by_ocr(pdf_path, page, outline, ocr_name):
    return extract_paras_by_outline(pdf_path, page, outline, ocr_name=ocr_name)


def remove_illegal_characters(characters: list):
    return [ILLEGAL_CHARACTERS_RE.sub(" ", x) if isinstance(x, str) else x for x in characters]


def dump_data_to_worksheet(workbook, header, data, sheet_name="sheet", sheet_index=0, worksheet=None):
    """
    header: ['a', 'b', 'c']
    data: [
        ['a1', 'b2', 'c3']
    ]
    ================================================================
    header: ['a', 'b', 'c']
    data: [
        {"a": [1]}, {"b": [1]}, {"c": [1]}
    ]
    """
    if not worksheet:
        worksheet = workbook.create_sheet(sheet_name, sheet_index)
    worksheet.append(header)
    if not isinstance(data[0], dict):
        for row_data in data:
            worksheet.append(remove_illegal_characters(row_data))
    else:
        # 可能存在header和cell没有一一对应，利用dict将cell一一对应，同时将拿不到的值置为None
        for row_data in data:
            clean_data = []
            for head in header:
                cell = row_data.get(head)
                if not cell:
                    text = None
                elif isinstance(cell, str):
                    text = cell
                else:
                    text = cell.pop(0)
                clean_data.append(text)
            worksheet.append(remove_illegal_characters(clean_data))
    return worksheet


async def ready_for_annotate_notify(fid, name):
    domain = get_config("web.domain") or "scriber"
    url = f"http://{domain}/#/search?fileid={fid}"
    tail = get_config("notification.ready_for_annotate_notify_tail") or ""
    message_lines = [f"文档 {name}(id={fid}) 已预处理完毕，可以进行标注：", f"{url}"]
    if tail:
        message_lines.append(tail)
    await MMPoster.send("\n".join(message_lines))


def get_ocr_expire_msg(pdfinsight):
    pages = pdfinsight.data.get("pages") or {}
    ocr_expired_pages = [
        int(page)
        for page, data in pages.items()
        if data.get("statis", {}).get("ocr_error_msg", "") == "license expired!"
    ]

    if not ocr_expired_pages:
        ocr_expired_msg = ""
    elif len(ocr_expired_pages) == len(pages):
        ocr_expired_msg = "OCR服务已过期，无法解析文件，请联系管理员"
    else:
        ocr_expired_msg = f"OCR服务已过期，{'、'.join((str(x + 1) for x in ocr_expired_pages))}页无法解析，请联系管理员"

    msg = {
        "is_ocr_expired": bool(ocr_expired_pages),
        "ocr_expired_msg": ocr_expired_msg,
        "ocr_expired_pages": ocr_expired_pages,
    }

    return msg


def is_page_header(pdfinsight, page_header_patterns, element):
    if not page_header_patterns:
        return False
    top_3_element_index = [x.index for x in pdfinsight.element_dict[element["page"]][:3]]
    if element["index"] not in top_3_element_index:
        return False
    if page_header_patterns.nexts(clean_txt(element["text"])):
        return True

    return False


def fix_ele_type(pdfinsight, page_header_patterns, ele_type, elt):
    if ele_type == "PARAGRAPH":
        if is_page_header(pdfinsight, page_header_patterns, elt):
            return "PAGE_HEADER"
    return ele_type


@contextlib.contextmanager
def limit_numpy_threads():
    """临时限制 NumPy 线程数的上下文管理器"""
    # 保存原始环境变量值
    original_env = {}
    thread_vars = (
        "OMP_NUM_THREADS",
        "OPENBLAS_NUM_THREADS",
        "MKL_NUM_THREADS",
        "VECLIB_MAXIMUM_THREADS",
        "NUMEXPR_NUM_THREADS",
    )

    # 保存当前环境变量值
    for var in thread_vars:
        original_env[var] = os.environ.get(var)
        os.environ[var] = "1"

    try:
        yield
    finally:
        # 恢复原始环境变量值
        for var in thread_vars:
            if original_env[var] is None:
                os.environ.pop(var, None)
            else:
                os.environ[var] = original_env[var]


def init_magika():
    import onnxruntime as rt
    from magika import Magika
    from magika.types.prediction_mode import PredictionMode

    class _Magika(Magika):
        def _init_onnx_session(self) -> rt.InferenceSession:
            options = rt.SessionOptions()
            options.inter_op_num_threads = 1
            options.intra_op_num_threads = 1
            # Disable onnxruntime's parallel execution to avoid performance issue on some platforms.
            return rt.InferenceSession(self._model_path, sess_options=options, providers=["CPUExecutionProvider"])

    return _Magika(prediction_mode=PredictionMode.BEST_GUESS)


def simple_match_ext(ext: str, path: Path | str | bytes, *expected: str) -> bool:
    if not isinstance(path, bytes) and not os.path.exists(path):
        logger.warning(f"File not exists: {path}")
        return False
    if not match_ext(path, *expected):
        logger.warning(f"File extension not match: {ext}, maybe not a {expected} file as expected.")
    return any(ext.lstrip(".").lower() == e.lstrip(".").lower() for e in expected)


def match_ext(path: Path | str | bytes, *expected: str) -> bool:
    result = ""
    if sys.version_info > (3, 12):
        if isinstance(path, str):
            path = Path(path)

        with limit_numpy_threads():
            magic = init_magika()
            magika_result = magic.identify_bytes(path) if isinstance(path, bytes) else magic.identify_path(path)
        if magika_result.score >= 0.6:
            result = magika_result.output.label
    if not result:
        result = filetype.guess_extension(path) or ""
    return any(result.lower() == ext.lstrip(".").lower() for ext in expected)


def excel_row_iter(path, sheet_index=0, skip_rows=0, values_only=False) -> Iterator[tuple[Any]]:
    """Iterate rows in excel file(both xls and xlsx)"""
    date_mode = None
    if match_ext(path, "xlsx"):
        import openpyxl

        work_book = openpyxl.load_workbook(path)
        work_sheet = work_book.worksheets[sheet_index]
        rows = work_sheet.iter_rows(min_row=skip_rows + 1, values_only=values_only)
    else:
        import xlrd

        work_book = xlrd.open_workbook(path)
        work_sheet = work_book.sheet_by_index(sheet_index)
        date_mode = work_book.datemode
        rows = work_sheet.get_rows()
        try:
            for _ in range(skip_rows):
                next(rows)
        except StopIteration as exp:
            raise StopIteration(f"No more rows in {path}") from exp

    for row in rows:
        if date_mode is not None:
            # legacy code, for xlrd only
            for cell in row:
                cell.date_mode = date_mode
            if values_only:
                row = tuple(cell.value for cell in row)
        yield row


def validate_timestamp(data: int) -> bool:
    from marshmallow import ValidationError

    try:
        datetime.fromtimestamp(data)
    except Exception as exp:
        raise ValidationError("Invalid timestamp.") from exp
    return True


@functools.lru_cache(maxsize=128)
def get_key_path(key, sep="-"):
    try:
        return sep.join([item.split(":")[0] for item in json.loads(key)[1:]])
    except Exception:
        return None


def custom_async_open(file_path, mode, context=None):
    context = context if context else python_aio_asyncio.AsyncioContext()
    return async_open(file_path, mode, context=context)


def compact_dumps(data, **kwargs) -> str:
    """json.dumps with compact format"""
    kwargs.update(
        {
            "ensure_ascii": False,
            "separators": (",", ":"),
        }
    )
    return json.dumps(data, **kwargs)


def custom_decode(data: bytes) -> str:
    # 根据编码方式解码数据
    try:
        decoding = chardet.detect(data)
        if decoding["confidence"] > 0.75:
            return data.decode(decoding["encoding"])
        return fallback_decode(data)
    except UnicodeDecodeError:
        return fallback_decode(data)


def add_time_hierarchy(created_utc: int, value: str, parent="") -> str:
    if created_utc is None or str(created_utc) == "" or str(created_utc) == "null":
        raise CustomError("created_utc is null")
    upload_date = datetime.fromtimestamp(created_utc)
    return os.path.join(
        parent, str(upload_date.year), str(upload_date.month), str(upload_date.day), value[:2], value[2:]
    )


def is_consecutive(nums: list[int]) -> bool:
    """
    判断整数列表是否连续
    """
    if not nums:
        return True

    # 排序列表
    sorted_nums = sorted(nums)

    # 检查相邻元素的差是否为1
    for i in range(1, len(sorted_nums)):
        if sorted_nums[i] - sorted_nums[i - 1] != 1:
            return False

    return True
