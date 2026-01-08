"""
一次性脚本
【银河证券】字段内容拆分
https://gitpd.paodingai.com/cheftin/docs_scriber/-/issues/704
"""
from copy import deepcopy
from pathlib import Path

import openpyxl
from openpyxl.reader.excel import load_workbook

from remarkable.common.pattern import PatternCollection
from remarkable.common.util import clean_txt, dump_data_to_worksheet, index_in_space_string


def split_open_rules(text, clean_text):
    # pid 4854 5283
    pattern_with_origin = PatternCollection(
        [
            r"[\s。]+((?:\d[^\u4e00-\u9fa5])临时开放[日期]+).?[\r\n]+(?:当发生|本基金)",
            r"[\s。]+(临时开放[日期]+).?[\r\n]+(?:当发生|本基金)",
        ]
    )
    features = [
        r"((?:\d[^\u4e00-\u9fa5])临时开放[日期]+)",
        r"[。]((?:\d.)本基金存续期间的其它临时开放安排)",
        r"以下情况管理人可设置临时开放日，临时开放日仅允许赎回、不允许认购[^。]*?管理人认为有利于保护投资者权益的其他情形",  # 7485
    ]
    prefixes = [r"(?:\d.)", r""]
    uncompleted_features = [
        r"{prefix}临时开放[日期]+[(（【].*[】）)].每月可增设",  # 16992
        r"{prefix}在本基金存续期内任何一个工作日[^。]*?基金资产单位净值首次[^。]*?基金管理人应[^。]*?日内开放",
        r"{prefix}存续期[间内].?(?:本?资产|基金)?管理人有权增设(?:临时)?开放日",
        r"{prefix}存续期[间内].?(?:本?资产|基金)?管理人有权(?:设立|调整|设置?)临时开放日",
        r"{prefix}存续期内管理人有权(?:增设|设立|调整|设置?)开放日",
        r"{prefix}本基金存续期间的其它临时开放安排",
        r"{prefix}本基金存续期间.当发生以下情况时管理人可增设(?:临时)?开放日",
        r"{prefix}本基金存续期间.当发生以下情况时管理人可(?:设立|调整|设置?)临时开放日",
        r"{prefix}本基金份额净值触及特定点位时增设开放期",
        r"{prefix}本?基金管理人[^。]*?可根据[^。]*?需要对开放日进行增设",
        r"{prefix}本?基金管理人[^。]*?可根据[^。]*?增设[^。]*?(?:临时)?开放日",
        r"{prefix}本?基金管理人[^。]*?可根据[^。]*?(?:设立|调整|设置?)[^。]*?临时开放日",
        r"{prefix}每月增设.?[次临时]*开放[日期]+",
        r"{prefix}以下情况管理人应增设开放日",
        r"{prefix}除固定开放日外，管理人有权根据产品运作需要增设开放日",
        r"{prefix}在满足[^。]*?的前提下，资产管理人有权调整参与和退出开放日",
        r"{prefix}本?基金管理人可在本基金封闭期内根据基金运作需要设置申购日",
        r"{prefix}每月可?增设[^。]*?临时开放[日期]+",
        r"{prefix}(?:[当发生以下情况时]{4,}.?)?本?(?:资产|基金)?管理人.?(?:可|依据)[^。]*?(?:增设|设立|调整|设置?).?[次临时]*开放[日期]+",
        r"{prefix}临时开放[日期]+",
    ]
    for item in uncompleted_features:
        for prefix in prefixes:
            features.append(deepcopy(item).replace("{prefix}", prefix))

    second_patterns = PatternCollection(features)
    cannot_split = PatternCollection(
        [
            r"^无$",
            r"不设开放日",
        ]
    )
    have_to_match = PatternCollection([r"增设|临时开放[日期]"])
    ret_from_origin_text = common_split(
        text, text, pattern_with_origin, cannot_split, have_to_match, index_contain_space=False
    )
    if ret_from_origin_text[0] == "match":
        return ret_from_origin_text

    return common_split(text, clean_text, second_patterns, cannot_split, have_to_match)


def split_monitoring(text, clean_text):
    # 先提 预警止损机制
    prefix = r"(?:\d+[^%\u4e00-\u9fa5])?"
    second_patterns = PatternCollection(
        [
            r"预警线及止损条款",
            r"基金份额净值触及特定点位时的增设开放期安排",
            r"为保护全体委托人利益.*?设置预警线及止损线",
            r"若经基金管理人估算的基金资产参考单位净值连续.个工作日低于预警线",
            r"当基金管理人与基金托管人核对一致的基金份颇净值小于或等于.*?时.标的基金触及预警线",
            r"当日.{,4}基金份额净值小于或等于.*?本基金触及预警线",
            r"当T日计划份额净值低于或等于预警线",
            rf"{prefix}(?:本基金设置)?预警止损机制",
            rf"{prefix}本基金的预警线",
            rf"{prefix}本(?:私募)?(?:基金|计划)不设置预警",
            rf"{prefix}若T日收盘时.*?本?(?:私募)?(?:基金|计划).*?净值.*?预[警替]线",
            rf"{prefix}[自在若当本]+(?:私募|资管)?(?:基金|计划)存续期限?内[任何每]+一个(?:工作|交易|估值)日.*?净值.*?预[警替]线",
            rf"{prefix}(?:当|若经)基金管理人(?:下一工作日).*?(?:估算|核对)的结果显示.*?本?(?:私募)?(?:基金|计划).*?净值.*?预[警替]线",
        ]
    )
    cannot_split = PatternCollection(
        [
            r"^无$",
        ]
    )

    have_to_match = PatternCollection(
        [
            r"预警",
        ]
    )
    waring_ret = common_split(text, clean_text, second_patterns, cannot_split=cannot_split, have_to_match=have_to_match)
    waring_text = None
    if waring_ret[0] == "match":
        waring_text = waring_ret[-1]
        text = waring_ret[1]  # 用剩下的文本来找 托管人日常监督具体项目
        clean_text = clean_txt(text)

    second_patterns = PatternCollection(
        [
            r"由托管人开展的?日常监控的具体项目",
            r"托管人日常监督具体项目",
            r"托管人负责日常投资监督",
            r"托管人对除管理人自行监督控制条款以外的投资比例内容开展日常监督，即代表基金托管人对基金投资运作进行了日常投资监督",
        ]
    )

    cannot_split = PatternCollection(
        [
            r"^无$",
        ]
    )

    have_to_match = PatternCollection(
        [
            r"(?<!不进行)日常",
        ]
    )
    rate_limit = common_split(text, clean_text, second_patterns, cannot_split, have_to_match=have_to_match)
    rate_limit = list(rate_limit)
    if rate_limit[0] != "match" and waring_ret[0] == "match":
        rate_limit[0] = "match"

    rate_limit[1] = rate_limit[1] or text
    rate_limit[2] = rate_limit[2] or text
    rate_limit.append(waring_text)
    return rate_limit


def split_reward(text, clean_text):
    second_patterns = PatternCollection(
        [
            r"(?:[(（【]\d[】）)、.])?业绩报酬的?支付",
            r"基金管理人于业绩报酬计算日起.*?向基金托管人发送.*?指令.?由基金托管人.*?支付给基金管理人.?",
            r"资产管理人向资产托管人发送.*?指令.?由资产托管人从.*?支付",
            r"基金管理人向基金托管人发送.*?指令.?基金托管人.*?支付",
            r"资产管理人确认业绩报酬.*?无误.*?由资产管理人.*?发送.*?指令.*?支付",
            r"(?:基金|资产)管理人向.*?发送.*?指令.*?支付",
            r"基金合同生效后.*支付",
        ]
    )
    cannot_split = PatternCollection(
        [
            r"^无$",
            r"^0元$",
            r"无业绩报酬。?$",
            r"无业绩报酬的提取及支付方式$",
            r"不再?(?:提取|计算|抽取|计提|收取|设置|设|提)(管理人)?(业绩报酬)?[。；.]?$",
            r"母基金不收取业绩报酬，在子基金层面收取",
            r"计提业绩报酬比例为0",
            r"本基金不计提超额业绩报酬",
        ]
    )
    have_to_match = PatternCollection([r"支付"])
    return common_split(text, clean_text, second_patterns, cannot_split, have_to_match)


def split_management_fee(text, clean_text):
    second_patterns = PatternCollection(
        [
            r"每日计提，按季支付。",
            r"自本私募基金成立起",
            r"自基金成立日起算",
            r"\d+管理人来函.*管理费的支付",
            r"计算方[法式]如下",
        ]
    )

    cannot_split = PatternCollection(
        [
            r"^无$",
        ]
    )
    return common_split(text, clean_text, second_patterns, cannot_split, have_to_match=None)


def split_hosting_fee(text, clean_text):
    second_patterns = PatternCollection(
        [
            r"自本?(?:私募)?基金成立日?起",
            r"(?:基金)?托管费.*?每日计(?:提|算)",
            r"[自本]*合同生效后",
            r"基金托管人自基金成立日.*?起.*?托管费.*支付",
            r"合伙企业应在每个核算日之后的5个工作日内向托管人支付托管费",
            r"本?(?:协议|合同|基金)(?:生效之日|起始运作日|成立之日)后.*?内一次性支付.*?托管费",
        ]
    )

    cannot_split = PatternCollection(
        [
            r"^无$",
        ]
    )
    return common_split(text, clean_text, second_patterns, cannot_split, have_to_match=None)


def split_service_fee(text, clean_text):
    second_patterns = PatternCollection(
        [
            r"自本?(?:私募)?基金成立日?起",
            r"(?:基金)?基金服务费.*?每日计(?:提|算)",
            r"[自本]*合同生效后",
            r"基金托管人自基金成立日.*?起.*?基金服务费.*支付",
            r"合伙企业应在每个核算日之后的5个工作日内向托管人支付基金服务费",
            r"本?(?:协议|合同|基金)(?:生效之日|起始运作日|成立之日)后.*?内一次性支付.*?基金服务费",
            r"每年基金服务费应于计提日之后.*?工作日完成支付",
            r"每.*[内末].*?管理人向.*?发送.*?划款指令.*?支付",
        ]
    )

    cannot_split = PatternCollection([r"^无$", r"不收取基金服务费。?$"])
    return common_split(text, clean_text, second_patterns, cannot_split, have_to_match=None)


def common_split(text, clean_text, second_patterns, cannot_split, have_to_match, index_contain_space=True):
    """
    :param text:
    :param clean_text:
    :param second_patterns: 第二部分应满足的正则
    :param cannot_split:
    :param have_to_match:
    :param index_contain_space: 计算在原始text中包含空白字符的范围
    :return:
    """
    if not text:
        return "no_match", None, None

    if cannot_split and cannot_split.nexts(clean_text):
        return "only_one", text, None

    if have_to_match and not have_to_match.nexts(clean_text):
        return "only_one", text, None

    if second_patterns:
        if second_match := second_patterns.nexts(clean_text):
            if "dst" in second_match.groupdict():
                c_start, c_end = second_match.span("dst")
            else:
                c_start, c_end = second_match.regs[-1]
            if index_contain_space:
                sp_start, sp_end = index_in_space_string(text, (c_start, c_end))
            else:
                sp_start, _ = c_start, c_end
            return "match", text[:sp_start], text[sp_start:]

    return "no_match", None, None


handlers_map = {
    "开放规则": split_open_rules,
    "监控比率及限制": split_monitoring,
    "业绩报酬计算方式": split_reward,
    "管理费": split_management_fee,
    "托管费": split_hosting_fee,
    "基金服务费": split_service_fee,
}


def get_value_from_sheet_row(row, index):
    value = row[index].value
    if value:
        if isinstance(value, str):
            value = value.strip()
    return value


def save_to_excel(keyname, append_headers, data):
    width_map = {
        "D": 70,
        "F": 60,
        "G": 60,
    }
    headers = ["PKID", "PID", "KEYNAME", "KEYVALUE", "is_match"]
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    dump_data_to_worksheet(workbook, headers + append_headers, data, worksheet=worksheet)
    for idx, width in width_map.items():
        worksheet.column_dimensions[idx].width = width
    dump_filename = f"{keyname}.xlsx"
    workbook.save(dump_filename)


def split_data(xlsx_path, keyname, append_headers):
    handler_func = handlers_map[keyname]
    records_workbook = load_workbook(xlsx_path)
    records_sheet = records_workbook["Sheet1"]
    data = []
    only_one, match_count, no_match_count, unknown_match = 0, 0, 0, 0

    for row in records_sheet.iter_rows():
        name = get_value_from_sheet_row(row, 4)
        if name != keyname:
            continue
        pkid = get_value_from_sheet_row(row, 0)
        # if pkid != 5283:
        #     continue
        pid = get_value_from_sheet_row(row, 1)
        keyvalue = get_value_from_sheet_row(row, 5) or ""
        item = [pkid, pid, keyname, keyvalue]
        ret = handler_func(keyvalue, clean_txt(keyvalue))
        if ret[0] == "only_one":
            only_one += 1
        elif ret[0] == "match":
            match_count += 1
        elif ret[0] == "unknown_match":
            unknown_match += 1
        else:
            no_match_count += 1

        item.extend(ret)
        # for i in ret:
        #     print('-------------')
        #     print(i)
        # return
        data.append(item)

    print(f"only_one: {only_one}, match_count: {match_count}")
    print(f"unknown_match: {unknown_match}, no_match_count: {no_match_count}")
    save_to_excel(keyname, append_headers, data)


def get_output(fields_map):
    history_path = Path("/Users/liuchao/Downloads/Scriber/china_stock/split_history/history_0")
    history_items = {}
    for xlsx_path in history_path.iterdir():
        if not xlsx_path.name.endswith("xlsx"):
            continue
        print(xlsx_path)
        records_workbook = load_workbook(xlsx_path)
        records_sheet = records_workbook["Sheet1"]

        for row in records_sheet.iter_rows(min_row=2):
            pkid = get_value_from_sheet_row(row, 0)
            new_item = [x.value for x in row]
            history_items[pkid] = new_item

    splited_path = Path("/Users/liuchao/Downloads/Scriber/china_stock/split_history/splited_0")
    output = []
    for xlsx_path in splited_path.iterdir():
        if not xlsx_path.name.endswith("xlsx"):
            continue
        print(xlsx_path)
        append_fields = None
        for keyname, append_headers in fields_map:
            if keyname in xlsx_path.name:
                append_fields = append_headers
                break

        records_workbook = load_workbook(xlsx_path)
        records_sheet = records_workbook["Sheet1"]

        for row in records_sheet.iter_rows(min_row=2):
            pkid = get_value_from_sheet_row(row, 0)
            for index, append_field in enumerate(append_fields):
                history_item = deepcopy(history_items[pkid])

                history_item[4] = append_field
                history_item[5] = get_value_from_sheet_row(row, 5 + index)
                output.append(history_item)
    headers = ["PKID", "PID", "PNAME", "KEYID", "KEYNAME", "KEYVALUE", "VALUESOURCE", "MID"]
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(index=0)
    dump_data_to_worksheet(workbook, headers, output, worksheet=worksheet)

    dump_filename = "output.xlsx"
    workbook.save(dump_filename)
    return history_items


def main():
    params = [
        ("开放规则", ["开放日", "临时开放日"]),
        ("监控比率及限制", ["投资比例限制", "由托管人开展日常监控的具体项目", "预警止损机制"]),
        ("业绩报酬计算方式", ["业绩报酬计算方式", "业绩报酬支付方式"]),
        ("管理费", ["管理费计算方式", "管理费支付方式"]),
        ("托管费", ["托管费计算方式", "托管费支付方式"]),
        ("基金服务费", ["基金服务费计算方式", "基金服务费支付方式"]),
    ]

    # handler_func = handlers_map[keyname]
    # data = handler_func('A临时开放日B')

    keyname, append_headers = params[2]
    xlsx_path = f"/Users/liuchao/Downloads/Scriber/china_stock/split_history/{keyname}_history.xlsx"
    split_data(xlsx_path, keyname, append_headers=append_headers)
    # get_output(params)


if __name__ == "__main__":
    main()
