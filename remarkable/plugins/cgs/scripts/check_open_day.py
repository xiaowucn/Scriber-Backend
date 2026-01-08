import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from remarkable.config import project_root
from remarkable.predictor.china_stock_predictor import PrivateFundContract


@dataclass
class FakeResult:
    text: str


@dataclass
class FakeSchema:
    name: str


def check():
    enum_predictor = PrivateFundContract()
    open_frequency = ['每日', '每周', '每月', '每季度', '每半年', '每年']
    data_path = Path(project_root) / 'data/tmp/china_stock/开放频率备注字段示例.xlsx'
    workbook = openpyxl.load_workbook(data_path, read_only=True)
    # workbook = openpyxl.load_workbook(data_path, data_only=True)

    schema = FakeSchema('开放日')
    total = 0
    wrong = 0
    for sheet_name in open_frequency:
        sheet = workbook.get_sheet_by_name(sheet_name)
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            text = row[0]
            frequency = row[1]

            if not text:
                break

            total += 1
            result = FakeResult(text)
            result = PrivateFundContract.clean_open_day(result)
            data = enum_predictor.predict(result, schema)
            # sheet.cell(idx + 1, 3).value = data
            if data != frequency:
                wrong += 1
                print(f'row: {idx + 2}, right: {frequency}, result: {data}')
                print(f'text: {text}')

    logging.info(f'{wrong=}')
    logging.info(f'{wrong/total=}')
    # workbook.save(data_path)


if __name__ == '__main__':
    check()
