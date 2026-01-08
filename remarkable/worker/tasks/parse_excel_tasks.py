import logging
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import oracledb
import xlrd

from remarkable.common.constants import PDFParseStatus
from remarkable.common.storage import localstorage
from remarkable.common.util import get_config, loop_wrapper, simple_match_ext
from remarkable.converter.gffunds import GFFundsWorkShop
from remarkable.converter.gffunds.table_struct import (
    TExcelParsingResult,
    TExcelParsingResultEx,
)
from remarkable.models.new_file import NewFile
from remarkable.worker.app import app

logger = logging.getLogger(__name__)


def parse_xls(path: Path) -> dict[int, list[dict]]:
    def _is_merged_cell(_row_idx, _col_idx):
        return any(r[0] <= _row_idx < r[1] and r[2] <= _col_idx < r[3] for r in merged_cells)

    workbook = xlrd.open_workbook(path, formatting_info=True)
    sheets_info = {}
    for sheet_idx, sheet in enumerate(workbook.sheets()):
        data = []
        merged_cells = sheet.merged_cells  # 起始行、结束行、起始列和结束列的索引
        for row_idx, row_cells in enumerate(sheet.get_rows()):
            row_dict = {}
            tb_merge = []
            for col_idx, cell in enumerate(row_cells):
                if _is_merged_cell(row_idx, col_idx):
                    tb_merge.append(f"{row_idx}_{col_idx}")
                row_dict.update({f"col{col_idx}": str(cell.value)})
            row_dict.update({"l_table_line": row_idx, "tb_merge": ",".join(tb_merge), "vc_sheet_name": sheet.name})
            data.append(row_dict)
        sheets_info.update({sheet_idx: data})
    return sheets_info


def parse_xlsx(path: Path) -> dict[int, list[dict]]:
    def _is_merged_cell(_row_idx, _col_idx):
        return any(
            merged_range.min_row <= _row_idx <= merged_range.max_row
            and merged_range.min_col <= _col_idx <= merged_range.max_col
            for merged_range in merged_cells
        )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=get_config("web.tmp_dir")) as temp_file:
        temp_file_path = temp_file.name
        shutil.copy2(path, temp_file_path)
        workbook = openpyxl.load_workbook(temp_file_path)
        sheets = workbook.worksheets
        sheets_info = {}
        for sheet_idx, sheet in enumerate(sheets):
            data = []
            merged_cells = sheet.merged_cells.ranges
            for idx, row in enumerate(sheet.iter_rows()):
                row_dict = {}
                tb_merge = []
                for cell in row:
                    row_idx = cell.row
                    col_idx = cell.column
                    if _is_merged_cell(row_idx, col_idx):
                        tb_merge.append(f"{row_idx - 1}_{col_idx - 1}")
                    row_dict.update({f"col{col_idx - 1}": str(cell.value)})
                row_dict.update({"l_table_line": idx, "tb_merge": ",".join(tb_merge), "vc_sheet_name": sheet.title})
                data.append(row_dict)
            sheets_info.update({sheet_idx: data})

        return sheets_info


def construct_table_dict(excel_data: dict[int, list[dict]], base_data: dict) -> tuple[list[dict], list[dict]]:
    results = []
    results_ex = []
    for sheet_idx, sheet_data in excel_data.items():
        for data in sheet_data:
            data.update(**base_data)
            data.update({"l_sheet_no": sheet_idx})
            parse_res = TExcelParsingResult(**data)
            if parse_res.is_clob is True:
                results_ex.append(
                    {
                        f"{base_data['vc_seq_no']}-{base_data['file_id']}-{data['l_sheet_no']}-{data['l_table_line']}": {
                            "dt_insert_time": base_data["dt_insert_time"],
                            "dt_update_time": base_data["dt_update_time"],
                            **parse_res.process_table_ex_attr(),
                        }
                    }
                )
            results.append(parse_res.to_dict())
    return results, results_ex


def insert_data_to_database(cursor, table_data: list[dict], table_data_ex: list[dict]):
    insert_sql = TExcelParsingResult.insert_sql()
    cursor.executemany(insert_sql, table_data)
    if table_data_ex:
        ex_data = []
        for data in table_data_ex:
            for key, value in data.items():
                vc_seq_no, file_id, l_sheet_no, l_table_line = key.split("-")
                query = TExcelParsingResult.query_row_sql(
                    vc_seq_no=vc_seq_no, file_id=file_id, l_sheet_no=l_sheet_no, l_table_line=l_table_line
                )
                row = cursor.execute(query)
                value.update({"vc_id": row.fetchone()[0]})
                ex_data.append(TExcelParsingResultEx(**value).to_dict())
        insert_ex_sql = TExcelParsingResultEx.insert_sql()
        cursor.executemany(insert_ex_sql, ex_data)


@app.task
@loop_wrapper
async def parse_excel_task(file_id: int, project_id: int, project_name: str):
    """【广发基金】公募月报（Excel）解析入库：docs_scriber#1778"""
    file = await NewFile.find_by_kwargs(id=file_id)
    path = localstorage.mount(file.path())
    if simple_match_ext(file.ext, path, "xls"):
        data = parse_xls(path)
    elif simple_match_ext(file.ext, path, "xlsx"):
        data = parse_xlsx(path)
    else:
        await file.update_(pdf_parse_status=PDFParseStatus.UNSUPPORTED_FILE)
        logger.error(f"文件{file.id=}: 该文件格式不在[xls,xlsx]中，解析失败")
        return
    current_time = datetime.now()
    base_info = {
        "file_id": file.id,
        "vc_seq_no": project_id,
        "project_name": project_name,
        "vc_report_type": "excel解析",
        "vc_file_name": file.name,
        "dt_insert_time": current_time,
        "dt_update_time": current_time,
    }
    table_data, table_data_ex = construct_table_dict(data, base_info)
    connection = GFFundsWorkShop._db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(TExcelParsingResult.delete_with_fid(file.id))
            insert_data_to_database(cursor, table_data, table_data_ex)
            connection.commit()
    except oracledb.Error as error:
        connection.rollback()
        logger.error(f"{file.id=} EXCEL解析入库失败{error}")
        await file.update_(pdf_parse_status=PDFParseStatus.EXCEL_INSERT_DB_FAILED)
        return
    logger.info(f"{file.id=} EXCEL解析入库成功")
    await file.update_(pdf_parse_status=PDFParseStatus.EXCEL_INSERT_DB_SUCCESS)
