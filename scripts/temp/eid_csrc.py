import calendar
import json
import re
import datetime
from collections import defaultdict
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

from remarkable.common.util import clean_txt


class EidCsrcStat:
    p_page = re.compile(r"共<b>(\d+)</b>页")

    szse_chuangyeban = {
        "exchange": "深交所",
        "start": "2020-05-01",
        "end": "2021-10-24",
        "temp_url": "http://eid.csrc.gov.cn/ipo/101012/index_{page}_f.html",
        "board_codes": {
            "申报稿": "03",
            "上会稿": "04",
            "注册稿": "05",
        },
        "sel_catagory": "10013",
    }

    szse_zhuban = {
        "exchange": "深交所",
        "start": "2020-09-19",
        "end": "2021-10-24",
        "temp_url": "http://eid.csrc.gov.cn/ipo/101011/index_{page}_f.html",
        "board_codes": {
            "预披露&预披露更新": "",
        },
        "sel_catagory": "9973",
    }

    sse_zhuban = {
        "exchange": "上交所",
        "start": "2020-09-19",
        "end": "2021-10-24",
        "temp_url": "http://eid.csrc.gov.cn/ipo/101010/index_{page}_f.html",
        "board_codes": {
            "预披露&预披露更新": "",
        },
        "sel_catagory": "9601",
    }

    headers = {
        "Host": "eid.csrc.gov.cn",
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer": "http://eid.csrc.gov.cn/ipo/101012/index.html",
        "Origin": "http://www.csrc.gov.cn",
    }

    time_format = "%Y-%m-%d"
    p_url = re.compile(r"http.*?pdf")
    mnt_url = re.compile(r"mnt/storage.*?pdf")

    @classmethod
    def stat_csrc_eid(cls, exchange, start, end, board_codes, temp_url, sel_catagory):
        for board, code in board_codes.items():
            cls.stat_board(exchange, board, code, start, end, temp_url, sel_catagory)

        excel_fields = ["name", "issue_type", "board", "institution", "date", "notice", "onclick"]
        for board, code in board_codes.items():
            json_file = f"{exchange}_{board}_{start}-{end}.json"
            with open(json_file) as file_obj:
                data = json.load(file_obj)
                filtered_data = cls.data_filter(data, cls.time_format)
                for date in data:
                    if date in filtered_data:
                        data[date] = filtered_data[date]
                    else:
                        data[date] = {"companies": [], "count": 0}

                # cls.download(exchange, board, data)
                cls.write_excel(exchange, start, end, board, data, excel_fields)

    @staticmethod
    def parse(text):
        ipo_companies = []
        soup = BeautifulSoup(text, "lxml")
        table = soup.find("table", class_="m-table2 m-table2-0")
        trs = table.find_all("tr")
        for tr in trs:
            if "招股说明书" in tr.text:
                tds = tr.find_all("td")
                ipo_companies.append(
                    {
                        "name": clean_txt(tds[0].text),
                        "issue_type": clean_txt(tds[1].text),
                        "board": clean_txt(tds[2].text),
                        "institution": clean_txt(tds[3].text),
                        "date": clean_txt(tds[4].text),
                        "notice": clean_txt(tds[5].text),
                        "onclick": clean_txt(tr.attrs["onclick"]),
                    }
                )
        return ipo_companies

    @classmethod
    def get_response(cls, temp_url, sel_catagory, page, board_code, start_date, end_date):
        form = {
            "selBoardCode": board_code,
            "selCatagory2": sel_catagory,
            "startDate": start_date,
            "endDate": end_date,
        }
        body = "&".join("{}={}".format(key, val) for key, val in form.items())
        url = temp_url.format(page=page)
        print(url)
        return requests.post(url, headers=cls.headers, params=body)

    @staticmethod
    def get_time_points(from_date, to_date, time_format):
        time_points = []
        start = datetime.datetime.strptime(from_date, time_format)
        end = datetime.datetime.strptime(to_date, time_format)

        date_point = start
        while True:
            if date_point > end:
                break
            date_point_str = datetime.datetime.strftime(date_point, time_format)
            time_points.append((date_point_str, date_point_str))
            date_point = date_point + datetime.timedelta(days=1)

        return time_points

    @staticmethod
    def get_month_points(from_date, to_date, time_format):
        time_points = []
        start = datetime.datetime.strptime(from_date, time_format)
        end = datetime.datetime.strptime(to_date, time_format)

        month_start = start
        while True:
            year, month = month_start.year, month_start.month
            _, month_days = calendar.monthrange(year, month)
            month_end = datetime.datetime(year, month, month_days)
            if month_end >= end:
                month_end = end

            time_points.append(
                (
                    datetime.datetime.strftime(month_start, time_format),
                    datetime.datetime.strftime(month_end, time_format),
                )
            )

            if month_end >= end:
                break
            else:
                month_start = month_end + datetime.timedelta(days=1)

        return time_points

    @classmethod
    def stat_board(cls, exchange, board, code, start, end, temp_url, sel_catagory):
        stat = defaultdict(list)
        if board == "预披露&预披露更新":
            date_points = cls.get_month_points(start, end, cls.time_format)
        else:
            date_points = cls.get_time_points(start, end, cls.time_format)
        for date_start, date_end in date_points:
            print(f"{date_start}-{date_end}")
            year_month = date_start[:7]
            response = cls.get_response(temp_url, sel_catagory, 1, code, date_start, date_end)
            ipo_companies = cls.parse(response.text)
            stat[year_month].extend(ipo_companies)

            match = cls.p_page.search(response.text)
            if match:
                pages = int(match.groups(0)[0])
                if pages > 50:
                    raise Exception("页码最多显示50页,数据可能不全")
                for page in range(2, pages + 1):
                    response = cls.get_response(temp_url, sel_catagory, page, code, date_start, date_end)
                    ipo_companies = cls.parse(response.text)
                    stat[year_month].extend(ipo_companies)
        board_stat = {k: {"companies": v, "count": len(v)} for k, v in stat.items()}
        with open(f"{exchange}_{board}_{start}-{end}.json", "w") as file_obj:
            json.dump(board_stat, file_obj, ensure_ascii=False)

    @staticmethod
    def write_excel(exchange, start, end, board, board_data, fields):
        work_book = Workbook()
        stat_sheet = work_book.get_sheet_by_name("Sheet")

        for index, date in enumerate(board_data.keys()):
            stat_sheet.cell(index + 1, 1).value = date
            stat_sheet.cell(index + 1, 2).value = board_data[date]["count"]

            sheet = work_book.create_sheet(date)
            for col_, field in enumerate(fields):
                sheet.cell(1, col_ + 1).value = field

            for _row, ans in enumerate(board_data[date]["companies"]):
                row = _row + 2
                for _col, key in enumerate(fields):
                    sheet.cell(row, _col + 1).value = ans.get(key)

        work_book.save(filename=f"{exchange}_{board}_{start}-{end}.xlsx")

    @staticmethod
    def data_filter(data, time_format):
        ret = defaultdict(list)
        filtered = defaultdict(dict)
        for value in data.values():
            for company in value["companies"]:
                name = company["name"]
                if name not in filtered:
                    filtered[name] = company
                else:
                    exist_date = datetime.datetime.strptime(filtered[name]["date"], time_format)
                    date = datetime.datetime.strptime(company["date"], time_format)
                    if date > exist_date:
                        filtered[name] = company

        for company in filtered.values():
            ret[company["date"][:7]].append(company)

        ret = {k: {"companies": v, "count": len(v)} for k, v in ret.items()}
        return ret

    @classmethod
    def download(cls, exchange, board, data):
        for date, value in data.items():
            download_dir = Path(f"{exchange}/{board}") / date
            if not download_dir.exists():
                download_dir.mkdir(parents=True)
            for company in value["companies"]:
                name = company["name"]
                file_name = cls.get_file_name(board, name)
                file_path = download_dir / f"{file_name}.pdf"
                if file_path.exists():
                    print(f'exists: {name}, {company["date"]}')
                    continue
                download_url = cls.get_download_url(company)
                print(f'download: {name}, {company["date"]}')
                response = requests.get(download_url)
                with open(file_path, "wb") as pdf:
                    pdf.write(response.content)

    @classmethod
    def get_download_url(cls, company):
        match = cls.p_url.search(company["onclick"])
        if match:
            return match.group()
        match = cls.mnt_url.search(company["onclick"])
        if match:
            mnt_path = match.group()
            return f"http://eid.csrc.gov.cn/{mnt_path}"

        return None

    @staticmethod
    def get_file_name(board, name):
        if board == "预披露&预披露更新":
            file_name = f"{name}首次公开发行股票招股说明书（申报稿）"
        else:
            file_name = f"{name}首次公开发行股票并在创业板上市招股说明书（{board}）"

        return file_name


if __name__ == "__main__":
    EidCsrcStat.stat_csrc_eid(**EidCsrcStat.sse_zhuban)
